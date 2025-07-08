import pandas as pd
from datetime import datetime
import os
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import telebot
from telebot import types
import io
import logging
import glob


# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize bot with your token
BOT_TOKEN ="7755249776:AAGwl7SjdxGEZnxGamostqLxcl-fz5Lq0mo"
bot = telebot.TeleBot(BOT_TOKEN)

ADMIN_LOG_FILE = "/home/nedquad12/user_activity.xlsx"
def log_user_activity(user_id, username, command):
    """Log user activity for admin tracking"""
    try:
        # Load existing data or create new
        if os.path.exists(ADMIN_LOG_FILE):
            df = pd.read_excel(ADMIN_LOG_FILE)
        else:
            df = pd.DataFrame(columns=['user_id', 'username', 'total_requests', 'last_command', 'last_activity'])
        
        # Check if user exists
        if user_id in df['user_id'].values:
            # Update existing user
            idx = df[df['user_id'] == user_id].index[0]
            df.loc[idx, 'total_requests'] += 1
            df.loc[idx, 'last_command'] = command
            df.loc[idx, 'last_activity'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            df.loc[idx, 'username'] = username  # Update username in case it changed
        else:
            # Add new user
            new_row = pd.DataFrame({
                'user_id': [user_id],
                'username': [username],
                'total_requests': [1],
                'last_command': [command],
                'last_activity': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            })
            df = pd.concat([df, new_row], ignore_index=True)
        
        # Save to Excel
        df.to_excel(ADMIN_LOG_FILE, index=False)
        
    except Exception as e:
        logger.error(f"Error logging user activity: {e}")

class TelegramStockDataViewer:
    def __init__(self, data_folder=None):
         # Automatically set folder path
        if data_folder:
            self.data_folder = data_folder
        else:
            self.data_folder = "/home/nedquad12/uyuu/data"  # Default relative folder
            
        self.margin_folder = "/home/nedquad12/uyuu/margin"
        self.margin_df = None
        self.margin_fields = ['Volume', 'Nilai', 'Frekuensi']
        
        # BlackRock folders
        self.blackrock_folders = {
            'indonesia': "/home/nedquad12/uyuu/br/ind",
            'china': "/home/nedquad12/uyuu/br/cn", 
            'us': "/home/nedquad12/uyuu/br/us",
            'btc': "/home/nedquad12/uyuu/br/btc"
        }
        
        # Data storage
        self.combined_df = None
        self.user_data = {}  # Store user-specific data (like chart selections)
        
        # BlackRock data storage
        self.blackrock_data = {
            'indonesia': None,
            'china': None,
            'us': None,
            'btc': None
        }

        # Available fields for plotting
        self.plot_fields = [
            'Local IS', 'Local CP', 'Local PF', 'Local IB', 'Local ID', 'Local MF', 'Local SC',
            'Foreign IS', 'Foreign CP', 'Foreign PF', 'Foreign IB', 'Foreign ID', 'Foreign MF', 'Foreign SC'
        ]
        
        # Nama cantik untuk tombol Telegram
        self.button_labels = {
            'Local IS': '🇮🇩 Lokal Asuransi',
            'Local CP': '🇮🇩 Lokal Korporat',
            'Local PF': '🇮🇩 Lokal Dana Pensiun',
            'Local IB': '🇮🇩 Lokal Bank',
            'Local ID': '🇮🇩 Lokal Ritel',
            'Local MF': '🇮🇩 Lokal Reksadana',
            'Local SC': '🇮🇩 Lokal Sekuritas',
            'Foreign IS': '🌏 Asing Asuransi',
            'Foreign CP': '🌏 Asing Korporat',
            'Foreign PF': '🌏 Asing Dana Pensiun',
            'Foreign IB': '🌏 Asing Bank',
            'Foreign ID': '🌏 Asing Ritel',
            'Foreign MF': '🌏 Asing Reksadana',
            'Foreign SC': '🌏 Asing Sekuritas'
             }


        # Auto-load data on initialization
        self.load_all_excel_files()
        self.load_margin_files()
        self.load_blackrock_data()
    
    def load_all_excel_files(self):
        """Load all Excel files from the data folder"""
        try:
            # Create data folder if it doesn't exist
            if not os.path.exists(self.data_folder):
                os.makedirs(self.data_folder)
                logger.info(f"📂 Created data folder: {self.data_folder}")
                return

            # Find all Excel files in the data folder
            excel_files = []
            for extension in ['*.xlsx', '*.xls', '*.XLSX', '*.XLS']:
                excel_files.extend(glob.glob(os.path.join(self.data_folder, extension)))

            if not excel_files:
                logger.warning("⚠️ No Excel files found in data folder")
                return

            logger.info(f"📄 Found {len(excel_files)} Excel files: {[os.path.basename(f) for f in excel_files]}")

            # Load all Excel files
            dataframes = []
            loaded_files = []

            for file_path in excel_files:
                try:
                    logger.info(f"📥 Loading file: {file_path}")
                    df = pd.read_excel(file_path)  # FIX: Use correct file path
                    dataframes.append(df)
                    loaded_files.append(os.path.basename(file_path))
                    logger.info(f"✅ Loaded: {os.path.basename(file_path)} - {len(df)} records")
                except Exception as e:
                    logger.error(f"❌ Error loading {file_path}: {e}")
                    continue

            if dataframes:
                # Combine all dataframes
                self.combined_df = self.combine_dataframes(dataframes)
                logger.info(f"✅ Successfully loaded {len(loaded_files)} files with {len(self.combined_df)} total records")
                logger.info(f"📅 Date range: {self.combined_df['Date'].min()} to {self.combined_df['Date'].max()}")
            else:
                logger.error("❌ No valid Excel files could be loaded")
                
            self.margin_df = None  

        except Exception as e:
            logger.error(f"❌ Error during auto-load: {e}")
            
    def load_margin_files(self):
        """Load margin trading files from margin folder"""
        try:
            if not os.path.exists(self.margin_folder):
                os.makedirs(self.margin_folder)
                logger.info(f"📂 Created margin folder: {self.margin_folder}")
                return

        # Find Excel files with ddmmyy.xlsx pattern
            excel_files = []
            for extension in ['*.xlsx', '*.xls']:
                excel_files.extend(glob.glob(os.path.join(self.margin_folder, extension)))

            if not excel_files:
               logger.warning("⚠️ No margin Excel files found")
               return

        # Limit to 60 files and sort by date
            excel_files = sorted(excel_files)[:60]
            logger.info(f"📄 Found {len(excel_files)} margin files")

            dataframes = []
            for file_path in excel_files:
                try:
                # Extract date from filename (ddmmyy.xlsx)
                    filename = os.path.basename(file_path)
                    date_str = filename.split('.')[0]
                
                    df = pd.read_excel(file_path)
                
                # Add date column based on filename
                    if len(date_str) == 6:  # ddmmyy format
                        day = int(date_str[:2])
                        month = int(date_str[2:4])
                        year = int('20' + date_str[4:6])  # assume 20xx
                        file_date = datetime(year, month, day)
                        df['Date'] = file_date
                
                    dataframes.append(df)
                    logger.info(f"✅ Loaded margin file: {filename}")
                
                except Exception as e:
                    logger.error(f"❌ Error loading margin file {file_path}: {e}")
                    continue

            if dataframes:
                self.margin_df = pd.concat(dataframes, ignore_index=True)
                self.margin_df = self.margin_df.sort_values('Date', ascending=True)
                logger.info(f"✅ Loaded {len(self.margin_df)} margin records")

        except Exception as e:
            logger.error(f"❌ Error loading margin files: {e}")

    def load_blackrock_data(self):
        """Load BlackRock data from all folders"""
        for region, folder_path in self.blackrock_folders.items():
            try:
                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)
                    logger.info(f"📂 Created BlackRock folder: {folder_path}")
                    continue

                # Find Excel files
                excel_files = []
                for extension in ['*.xlsx', '*.xls']:
                    excel_files.extend(glob.glob(os.path.join(folder_path, extension)))

                if not excel_files:
                    logger.warning(f"⚠️ No BlackRock files found in {region}")
                    continue

                # Sort files by date (newest first) and limit to 60
                excel_files = sorted(excel_files, reverse=True)[:60]
                logger.info(f"📄 Found {len(excel_files)} BlackRock files for {region}")

                dataframes = []
                for file_path in excel_files:
                    try:
                        # Extract date from filename (ddmmyy.xlsx)
                        filename = os.path.basename(file_path)
                        date_str = filename.split('.')[0]
                        
                        df = pd.read_excel(file_path)
                        
                        # Add date column based on filename
                        if len(date_str) == 6:  # ddmmyy format
                            day = int(date_str[:2])
                            month = int(date_str[2:4])
                            year = int('20' + date_str[4:6])  # assume 20xx
                            file_date = datetime(year, month, day)
                            df['Date'] = file_date
                        
                        dataframes.append(df)
                        logger.info(f"✅ Loaded BlackRock file: {filename} for {region}")
                        
                    except Exception as e:
                        logger.error(f"❌ Error loading BlackRock file {file_path}: {e}")
                        continue

                if dataframes:
                    self.blackrock_data[region] = pd.concat(dataframes, ignore_index=True)
                    self.blackrock_data[region] = self.blackrock_data[region].sort_values('Date', ascending=True)
                    logger.info(f"✅ Loaded {len(self.blackrock_data[region])} BlackRock records for {region}")

            except Exception as e:
                logger.error(f"❌ Error loading BlackRock data for {region}: {e}")
    
    def reload_data(self):
        """Reload all data from the data folder"""
        self.combined_df = None
        self.margin_df = None
        self.load_all_excel_files()
        self.load_margin_files()
        self.load_blackrock_data()
    
    def get_user_data(self, user_id):
        if user_id not in self.user_data:
            self.user_data[user_id] = {
                'chart_selections': set()
            }
        return self.user_data[user_id]
    
    def load_excel_file(self, file_path):
        try:
            df = pd.read_excel(file_path)
            return df
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return None
    
    def combine_dataframes(self, dfs):
        try:
        # Combine all dataframes
           combined_df = pd.concat(dfs, ignore_index=True)
        
        # Convert Date column
           combined_df['Date'] = pd.to_datetime(combined_df['Date'])
        
        # Remove duplicates based on Date and Code
           combined_df = combined_df.drop_duplicates(subset=['Date', 'Code'], keep='last')
        
        # Sort by date (oldest first - ascending order)
           combined_df = combined_df.sort_values('Date', ascending=True)
        
           return combined_df
        except Exception as e:
           logger.error(f"Error combining dataframes: {e}")
        return None
    
    def search_margin_stock(self, code):
        """Search margin data for specific stock"""
        if self.margin_df is None:
            return None
    
        stock_data = self.margin_df[self.margin_df['Kode Saham'].str.upper() == code.upper()]
        return stock_data if not stock_data.empty else None
    
    def search_stock(self, code, limit=6):
        if self.combined_df is None:
           return None
    
    # Filter data by code and limit to specified records (oldest first)
        stock_data = self.combined_df[self.combined_df['Code'].str.upper() == code.upper()].head(limit)
    
        if stock_data.empty:
           return None
    
    # Debug: Print raw data to check for duplicates
        logger.info(f"Raw data for {code}:\n{stock_data[['Date', 'Code']]}")
    
        return stock_data
    
    def search_blackrock_ticker(self, region, ticker):
        """Search BlackRock data for specific ticker"""
        if region not in self.blackrock_data or self.blackrock_data[region] is None:
            return None
        
        ticker_data = self.blackrock_data[region][
            self.blackrock_data[region]['Ticker'].str.upper() == ticker.upper()
        ]
        return ticker_data if not ticker_data.empty else None
    
    def get_all_stock_codes(self):
        """Get all unique stock codes from the data"""
        if self.combined_df is None:
           return []

        if 'Code' not in self.combined_df.columns:
           logger.error("❌ Column 'Code' not found in data")
           return []

        try:
          # Pastikan hanya ambil nilai string
           codes = self.combined_df['Code'].dropna()
           codes = codes[codes.apply(lambda x: isinstance(x, str))]
           return sorted(codes.unique())
        except Exception as e:
           logger.error(f"❌ Error getting stock codes: {e}")
           return []
   
    def get_data_info(self):
        """Get information about the loaded data"""
        # Regular data info
        if self.combined_df is None:
            regular_info = "No regular data loaded"
        else:
            total_records = len(self.combined_df)
            unique_codes = len(self.combined_df['Code'].unique())
            date_range = f"{self.combined_df['Date'].min().strftime('%d-%b-%Y')} to {self.combined_df['Date'].max().strftime('%d-%b-%Y')}"
            regular_info = {
                'total_records': total_records,
                'unique_codes': unique_codes,
                'date_range': date_range
            }

        # Margin data info
        if self.margin_df is None:
            margin_info = "No margin data loaded"
        else:
            margin_records = len(self.margin_df)
            margin_codes = len(self.margin_df['Kode Saham'].unique())
            margin_range = f"{self.margin_df['Date'].min().strftime('%d-%b-%Y')} to {self.margin_df['Date'].max().    strftime('%d-%b-%Y')}"
            margin_info = {
                'total_records': margin_records,
                'unique_codes': margin_codes,
                'date_range': margin_range
            }

        return {
            'regular': regular_info,
            'margin': margin_info
        }   
    
    def format_stock_data(self, data):
        if data is None or data.empty:
            return "No data found"
        
        local_columns = ['Local IS', 'Local CP', 'Local PF', 'Local IB', 'Local ID', 'Local MF', 'Local SC']
        foreign_columns = ['Foreign IS', 'Foreign CP', 'Foreign PF', 'Foreign IB', 'Foreign ID', 'Foreign MF', 'Foreign SC']
        
        result = []
        for _, row in data.iterrows():
            # Calculate totals
            local_total = sum(row.get(col, 0) for col in local_columns if pd.notna(row.get(col, 0)))
            foreign_total = sum(row.get(col, 0) for col in foreign_columns if pd.notna(row.get(col, 0)))
            
            result.append(f"📅 {row['Date'].strftime('%d-%b-%Y')}")
            result.append(f"🏷️ Code: {row.get('Code', '')}")
            result.append(f"📊 Type: {row.get('Type', '')}")
            result.append(f"💰 Price: {row.get('Price', 0):,.0f}")
            result.append(f"🏠 Total Local: {local_total:,.0f}")
            result.append(f"🌍 Total Foreign: {foreign_total:,.0f}")
            result.append("─" * 30)
        
        return "\n".join(result)
    
    def create_margin_charts(self, code):
        """Create 3 separate charts for Volume, Nilai, Frekuensi"""
        margin_data = self.search_margin_stock(code)
        if margin_data is None:
            return None
    
        grouped = margin_data.groupby('Date')[self.margin_fields].sum().sort_index()
    
        fig, axes = plt.subplots(3, 1, figsize=(12, 15))
    
    # Volume Chart
        axes[0].plot(grouped.index, grouped['Volume'], marker='o', color='blue', linewidth=2)
        axes[0].set_title(f'Volume - {code}', fontsize=14, fontweight='bold')
        axes[0].set_ylabel('Volume')
        axes[0].grid(True, alpha=0.3)
    
    # Nilai Chart
        axes[1].plot(grouped.index, grouped['Nilai'], marker='s', color='green', linewidth=2)
        axes[1].set_title(f'Nilai - {code}', fontsize=14, fontweight='bold')
        axes[1].set_ylabel('Nilai')
        axes[1].grid(True, alpha=0.3)
    
    # Frekuensi Chart
        axes[2].plot(grouped.index, grouped['Frekuensi'], marker='^', color='red', linewidth=2)
        axes[2].set_title(f'Frekuensi - {code}', fontsize=14, fontweight='bold')
        axes[2].set_ylabel('Frekuensi')
        axes[2].set_xlabel('Date')
        axes[2].grid(True, alpha=0.3)
    
        plt.tight_layout()
    
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=300, bbox_inches='tight')
        buf.seek(0)
        plt.close()
    
        return buf
    
    def create_excel_report(self, data, file_path, code):
        wb = Workbook()
        ws = wb.active
        ws.title = f"{code} Analysis"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center')
        
        # Headers
        headers = ['Date', 'Code', 'Type', 'Price', 'Local IS', 'Local CP', 'Local PF', 'Local IB', 
                  'Local ID', 'Local MF', 'Local SC', 'Foreign IS', 'Foreign CP', 'Foreign PF', 
                  'Foreign IB', 'Foreign ID', 'Foreign MF', 'Foreign SC']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        # Add Total columns headers
        ws.cell(row=1, column=len(headers)+1, value="Total Local").font = header_font
        ws.cell(row=1, column=len(headers)+1).fill = header_fill
        ws.cell(row=1, column=len(headers)+1).border = border
        ws.cell(row=1, column=len(headers)+1).alignment = center_align
        
        ws.cell(row=1, column=len(headers)+2, value="Total Foreign").font = header_font
        ws.cell(row=1, column=len(headers)+2).fill = header_fill
        ws.cell(row=1, column=len(headers)+2).border = border
        ws.cell(row=1, column=len(headers)+2).alignment = center_align
        
        # Write data
        local_cols = ['Local IS', 'Local CP', 'Local PF', 'Local IB', 'Local ID', 'Local MF', 'Local SC']
        foreign_cols = ['Foreign IS', 'Foreign CP', 'Foreign PF', 'Foreign IB', 'Foreign ID', 'Foreign MF', 'Foreign SC']
        
        for row_idx, (_, row_data) in enumerate(data.iterrows(), 2):
            # Basic data
            ws.cell(row=row_idx, column=1, value=row_data['Date'].strftime('%d-%b-%Y')).border = border
            ws.cell(row=row_idx, column=2, value=row_data.get('Code', '')).border = border
            ws.cell(row=row_idx, column=3, value=row_data.get('Type', '')).border = border
            ws.cell(row=row_idx, column=4, value=row_data.get('Price', 0)).border = border
            
            # Local and Foreign data
            for col_idx, header in enumerate(headers[4:], 5):
                value = row_data.get(header, 0)
                ws.cell(row=row_idx, column=col_idx, value=value).border = border
            
            # Add formulas for totals
            local_range = f"E{row_idx}:K{row_idx}"  # Local columns
            foreign_range = f"L{row_idx}:R{row_idx}"  # Foreign columns
            
            total_local_cell = ws.cell(row=row_idx, column=len(headers)+1)
            total_local_cell.value = f"=SUM({local_range})"
            total_local_cell.border = border
            
            total_foreign_cell = ws.cell(row=row_idx, column=len(headers)+2)
            total_foreign_cell.value = f"=SUM({foreign_range})"
            total_foreign_cell.border = border
        
        # Add summary section
        summary_row = len(data) + 3
        ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=14)
        ws.cell(row=summary_row+1, column=1, value=f"Stock Code: {code}")
        ws.cell(row=summary_row+2, column=1, value=f"Total Records: {len(data)}")
        ws.cell(row=summary_row+3, column=1, value=f"Date Range: {data['Date'].min().strftime('%d-%b-%Y')} to {data['Date'].max().strftime('%d-%b-%Y')}")
        
        # Add grand total formulas
        last_row = len(data) + 1
        ws.cell(row=summary_row+5, column=1, value="Grand Total Local:").font = Font(bold=True)
        ws.cell(row=summary_row+5, column=2, value=f"=SUM(S2:S{last_row})").font = Font(bold=True)
        ws.cell(row=summary_row+6, column=1, value="Grand Total Foreign:").font = Font(bold=True)
        ws.cell(row=summary_row+6, column=2, value=f"=SUM(T2:T{last_row})").font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    
    def create_line_chart(self, selected_fields, code=None):
        if self.combined_df is None:
            return None
        
        # Make a copy of the dataframe
        df = self.combined_df.copy()
        
        # Calculate totals if needed
        if 'Total Local' in selected_fields:
            df['Total Local'] = df[[f for f in self.plot_fields if f.startswith('Local')]].sum(axis=1)
        if 'Total Foreign' in selected_fields:
            df['Total Foreign'] = df[[f for f in self.plot_fields if f.startswith('Foreign')]].sum(axis=1)
        
        # Filter stock if applicable
        if code:
            df = df[df['Code'].str.upper() == code.upper()]
            if df.empty:
                return None
            
            
        
        # Group by date and sum
        grouped = df.groupby('Date')[selected_fields].sum().sort_index()
        
        plt.figure(figsize=(12, 8))
        for field in selected_fields:
            plt.plot(grouped.index, grouped[field], marker='o', label=field, linewidth=2)
        
        plt.xlabel('Date', fontsize=12)
        plt.ylabel('Value', fontsize=12)
        plt.title(f"Line Chart{' for ' + code if code else ''}", fontsize=14, fontweight='bold')
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        
        # Save to bytes
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=300, bbox_inches='tight')
        buf.seek(0)
        plt.close()
        
        return buf
    
    def create_blackrock_chart(self, region, ticker):
        """Create chart for BlackRock ticker data"""
        ticker_data = self.search_blackrock_ticker(region, ticker)
        if ticker_data is None:
            return None, None
        
        # Group by date and get latest values for each date
        grouped = ticker_data.groupby('Date').last().sort_index()
        
        # Create chart for Quantity Total
        plt.figure(figsize=(12, 8))
        plt.plot(grouped.index, grouped['Quantity Total'], marker='o', linewidth=2, color='blue')
        plt.title(f'BlackRock Holdings - {ticker} ({region.upper()})', fontsize=14, fontweight='bold')
        plt.xlabel('Date', fontsize=12)
        plt.ylabel('Quantity Total', fontsize=12)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        
        # Save chart to buffer
        chart_buf = io.BytesIO()
        plt.savefig(chart_buf, format='png', dpi=300, bbox_inches='tight')
        chart_buf.seek(0)
        plt.close()
        
        # Generate movement caption
        caption = self.generate_movement_caption(grouped, ticker, region)
        
        return chart_buf, caption

    def generate_movement_caption(self, grouped_data, ticker, region):
        """Generate detailed movement caption"""
        if len(grouped_data) < 2:
            return f"📊 BlackRock Holdings - {ticker} ({region.upper()})\n❌ Insufficient data for movement analysis"
        
        # Get latest and previous data
        latest = grouped_data.iloc[-1]
        previous = grouped_data.iloc[-2]
        
        # Calculate changes
        qty_change = latest['Quantity Total'] - previous['Quantity Total']
        qty_change_pct = (qty_change / previous['Quantity Total']) * 100 if previous['Quantity Total'] != 0 else 0
        
        mv_change = latest['Market Value Total'] - previous['Market Value Total']
        mv_change_pct = (mv_change / previous['Market Value Total']) * 100 if previous['Market Value Total'] != 0 else 0
        
        # Format numbers
        qty_latest = f"{latest['Quantity Total']:,.0f}"
        qty_prev = f"{previous['Quantity Total']:,.0f}"
        mv_latest = f"{latest['Market Value Total']:,.0f}"
        mv_prev = f"{previous['Market Value Total']:,.0f}"
        
        # Direction indicators
        qty_arrow = "🔺" if qty_change > 0 else "🔻" if qty_change < 0 else "➡️"
        mv_arrow = "🔺" if mv_change > 0 else "🔻" if mv_change < 0 else "➡️"
        
        caption = f"""📊 BlackRock Holdings - {ticker} ({region.upper()})

📅 Latest: {latest.name.strftime('%d-%b-%Y')}
📅 Previous: {previous.name.strftime('%d-%b-%Y')}

📈 Quantity Total:
Current: {qty_latest}
Previous: {qty_prev}
Change: {qty_arrow} {qty_change:+,.0f} ({qty_change_pct:+.2f}%)

💰 Market Value Total:
Current: ${mv_latest}
Previous: ${mv_prev}
Change: {mv_arrow} ${mv_change:+,.0f} ({mv_change_pct:+.2f}%)"""
        
        return caption

    def get_significant_movements(self, threshold=3.0):
        """Get all tickers with significant movements (>= threshold%)"""
        movements = []
        
        for region, data in self.blackrock_data.items():
            if data is None or len(data) < 2:
                # Limit to latest 5 dates for /b7 command only
                latest_dates = sorted(data['Date'].unique(), reverse=True)[:5]
                data = data[data['Date'].isin(latest_dates)]
                continue
                
            # Get unique tickers
            tickers = data['Ticker'].unique()
            
            for ticker in tickers:
                ticker_data = data[data['Ticker'] == ticker].groupby('Date').last().sort_index()
                
                if len(ticker_data) < 2:
                    continue
                    
                # Calculate movement
                latest = ticker_data.iloc[-1]
                previous = ticker_data.iloc[-2]
                
                if previous['Quantity Total'] == 0:
                    continue
                    
                qty_change_pct = ((latest['Quantity Total'] - previous['Quantity Total']) / previous['Quantity Total']) * 100
                
                if abs(qty_change_pct) >= threshold:
                    movements.append({
                        'region': region,
                        'ticker': ticker,
                        'change_pct': qty_change_pct,
                        'latest_qty': latest['Quantity Total'],
                        'previous_qty': previous['Quantity Total'],
                        'latest_mv': latest['Market Value Total'],
                        'previous_mv': previous['Market Value Total'],
                        'latest_date': latest.name,
                        'previous_date': previous.name
                    })
        
        # Sort by absolute change percentage
        movements.sort(key=lambda x: abs(x['change_pct']), reverse=True)
        return movements


# Initialize the viewer
viewer = TelegramStockDataViewer()

@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    data_info = viewer.get_data_info()
    
    if isinstance(data_info['regular'], str):
        data_status = "❌ No data loaded"
    else:
        regular = data_info['regular']
        data_status = f"✅ Data Holding loaded: {regular['total_records']} records, {regular['unique_codes']} stocks\n📅 Date range: {regular['date_range']}"
    
    # Perbaikan: Ubah ke format yang lebih sederhana tanpa parsing entities
    help_text = f"""🤖 Bot Analisa saha,

Perintah untuk IHSG
/help - Memanggil pesan ini
/search [CODE] - Search untuk data holding saham
/export [CODE] - Export data ke Excel
/chart - Membuat Chart
/reload - Reload data dari server
/m [CODE] - Menampikan data transaksi Margin

Perintah untuk Blacrock 
/bi for Indonesia stock
/bc for China stock
/btc for Bitcoin
/bu for US stock (S&P500)
/b7 - Pergerakan besar

Contoh:
- /search BBCA - Mencari data saham BBCA
- /bi BBCA - Mencari saham BBCA yang dimiliki Blackrock

Twitter Owner: https://x.com/saberial_link/
Telegram Owner: @Rendanggedang"""
    
    # Kirim tanpa parse_mode untuk menghindari error parsing
    bot.reply_to(message, help_text)

@bot.message_handler(commands=['reload'])
def reload_data(message):
    try:
        bot.reply_to(message, "🔄 Memuat ulang data dari server...")
        viewer.reload_data()
        
        data_info = viewer.get_data_info()
        if isinstance(data_info, str):
            bot.reply_to(message, "❌ Tidak ada data yang dimuat dari server.")
        else:
            response = f"✅ Data berhasil dimuat ulang!\n📊 Total pencatatan: {data_info['total_records']}\n📈 Jumlsh saham: {data_info['unique_codes']}"
            bot.reply_to(message, response)
    except Exception as e:
        bot.reply_to(message, f"❌ Error memuat data: {str(e)}")
    
@bot.message_handler(commands=['m'])
def margin_trading(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    if viewer.margin_df is None:
        bot.reply_to(message, "❌ No margin data loaded. Saham ini tidak terdaftar dalam margin.")
        return
    
    parts = message.text.split()
    if len(parts) < 2:
        bot.reply_to(message, "❌ Please provide a stock code.\nExample: `/m BBCA`", parse_mode='Markdown')
        return
    
    code = parts[1].upper()
    
    try:
        chart_buffer = viewer.create_margin_charts(code)
        if chart_buffer is None:
            bot.reply_to(message, f"❌ No margin data found for stock code: {code}")
            return
        
        # Send chart
        bot.send_photo(
            message.chat.id,
            chart_buffer,
            caption=f"📊 Transaction Margin for {code}\n📈 Volume, Nilai, Frekuensi"
        )
        
    except Exception as e:
        bot.reply_to(message, f"❌ Error creating margin chart: {str(e)}")
        logger.error(f"Error creating margin chart: {e}")

@bot.message_handler(commands=['export'])
def export_stock(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    if viewer.combined_df is None:
        bot.reply_to(message, "❌ No data loaded. Please check if Excel files are in the 'data' folder and use /reload.")
        return
    
    # Extract stock code from command
    parts = message.text.split()
    if len(parts) < 2:
        bot.reply_to(message, "❌ Please provide a stock code.\nExample: /export BBCA")
        return
    
    code = parts[1].upper()
    
    # Search for stock
    stock_data = viewer.search_stock(code)
    
    if stock_data is None:
        bot.reply_to(message, f"❌ No data found for stock code: {code}")
        return
    
    try:
        # Create Excel report
        export_file = f"export_{code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        viewer.create_excel_report(stock_data, export_file, code)
        
        # Send file
        with open(export_file, 'rb') as f:
            bot.send_document(message.chat.id, f, caption=f"📊 Stock analysis for {code}")
        
        # Clean up
        os.remove(export_file)
        
    except Exception as e:
        bot.reply_to(message, f"❌ Error creating Excel report: {str(e)}")
        logger.error(f"Error creating Excel report: {e}")

@bot.message_handler(commands=['bi'])
def blackrock_indonesia(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    handle_blackrock_command(message, 'indonesia')

@bot.message_handler(commands=['bc'])
def blackrock_china(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    handle_blackrock_command(message, 'china')

@bot.message_handler(commands=['bu'])
def blackrock_us(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    handle_blackrock_command(message, 'us')

@bot.message_handler(commands=['btc'])
def blackrock_btc(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    # Bitcoin only - tidak perlu ticker
    if viewer.blackrock_data['btc'] is None:
        bot.reply_to(message, "❌ No BlackRock Bitcoin data loaded.")
        return
    
    try:
        # Assume Bitcoin ticker for BTC
        chart_buffer, caption = viewer.create_blackrock_chart('btc', 'BTC')
        if chart_buffer is None:
            bot.reply_to(message, "❌ No Bitcoin data found")
            return
        
        bot.send_photo(message.chat.id, chart_buffer, caption=caption)
        
    except Exception as e:
        bot.reply_to(message, f"❌ Error creating BlackRock Bitcoin chart: {str(e)}")
        logger.error(f"Error creating BlackRock Bitcoin chart: {e}")

def handle_blackrock_command(message, region):
    """Handle BlackRock commands that require ticker input"""
    if viewer.blackrock_data[region] is None:
        bot.reply_to(message, f"❌ No BlackRock {region} data loaded.")
        return
    
    parts = message.text.split()
    if len(parts) < 2:
        bot.reply_to(message, f"❌ Please provide a ticker.\nExample: /{parts[0][1:]} AAPL")
        return
    
    ticker = parts[1].upper()
    
    try:
        chart_buffer, caption = viewer.create_blackrock_chart(region, ticker)
        if chart_buffer is None:
            bot.reply_to(message, f"❌ No data found for ticker: {ticker} in {region}")
            return
        
        bot.send_photo(message.chat.id, chart_buffer, caption=caption)
        
    except Exception as e:
        bot.reply_to(message, f"❌ Error creating BlackRock {region} chart: {str(e)}")
        logger.error(f"Error creating BlackRock {region} chart: {e}")

@bot.message_handler(commands=['b7'])
def blackrock_significant_movements(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    """Show significant BlackRock movements (>= 3%)"""
    try:
        movements = viewer.get_significant_movements(3.0)
        
        if not movements:
            bot.reply_to(message, "📊 No significant BlackRock movements (>= 3%) found in the last period.")
            return
        
        # Format response
        response = "📊 BlackRock Significant Movements (>= 3%)\n\n"
        
        for i, movement in enumerate(movements[:20]):  # Limit to top 20
            arrow = "🔺" if movement['change_pct'] > 0 else "🔻"
            response += f"{i+1}. {movement['ticker']} ({movement['region'].upper()})\n"
            response += f"   {arrow} {movement['change_pct']:+.2f}%\n"
            response += f"   Latest: {movement['latest_date'].strftime('%d-%b-%Y')}\n"
            response += f"   Previous: {movement['previous_date'].strftime('%d-%b-%Y')}\n"
            response += f"   Qty: {movement['latest_qty']:,.0f}\n"
            response += f"   MV: ${movement['latest_mv']:,.0f}\n\n"
        
        if len(movements) > 20:
            response += f"... and {len(movements) - 20} more movements"
        
        bot.reply_to(message, response)
        
    except Exception as e:
        bot.reply_to(message, f"❌ Error getting significant movements: {str(e)}")
        logger.error(f"Error getting significant movements: {e}")

@bot.message_handler(commands=['chart'])
def create_chart(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    if viewer.combined_df is None:
        bot.reply_to(message, "❌ No data loaded. Please check if Excel files are in the 'data' folder and use /reload.")
        return
    
    # Extract stock code from command (optional)
    parts = message.text.split()
    code = parts[1].upper() if len(parts) > 1 else None
    
    # Create keyboard for field selection
    markup = types.InlineKeyboardMarkup(row_width=2)
    
    # Add field selection buttons
    field_buttons = []
    for field in viewer.plot_fields:
        label = viewer.button_labels.get(field, field)  # Pakai nama cantik
        callback_data = f"field_{field.replace(' ', '_')}"
        field_buttons.append(types.InlineKeyboardButton(label, callback_data=callback_data))
    
    # Add buttons in pairs
    for i in range(0, len(field_buttons), 2):
        if i + 1 < len(field_buttons):
            markup.add(field_buttons[i], field_buttons[i + 1])
        else:
            markup.add(field_buttons[i])
    
    # Add total buttons
    markup.add(
        types.InlineKeyboardButton("Total Local", callback_data="field_Total_Local"),
        types.InlineKeyboardButton("Total Foreign", callback_data="field_Total_Foreign")
    )
    
    # Add control buttons
    markup.add(
        types.InlineKeyboardButton("Select All", callback_data="select_all"),
        types.InlineKeyboardButton("Clear All", callback_data="clear_all")
    )
    markup.add(types.InlineKeyboardButton("Generate Chart", callback_data=f"generate_chart_{code if code else 'all'}"))
    
    chart_text = f"📊 Chart Configuration\n"
    if code:
        chart_text += f"Stock: {code}\n"
    chart_text += "\nSelect fields to include in the chart:"
    
    # Kirim tanpa parse_mode
    bot.reply_to(message, chart_text, reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith('field_') or call.data in ['select_all', 'clear_all'] or call.data.startswith('generate_chart_'))

@bot.callback_query_handler(func=lambda call: call.data.startswith('field_') or call.data in ['select_all', 'clear_all'] or call.data.startswith('generate_chart_'))
def handle_chart_selection(call):
    user_id = call.from_user.id
    user_data = viewer.get_user_data(user_id)

    if call.data.startswith('field_'):
        field = call.data[6:].replace('_', ' ')
        if field in user_data['chart_selections']:
            user_data['chart_selections'].remove(field)
        else:
            user_data['chart_selections'].add(field)

    elif call.data == 'select_all':
        user_data['chart_selections'] = set(viewer.plot_fields + ['Total Local', 'Total Foreign'])

    elif call.data == 'clear_all':
        user_data['chart_selections'] = set()

    elif call.data.startswith('generate_chart_'):
        code = call.data[15:] if call.data[15:] != 'all' else None

        if not user_data['chart_selections']:
            bot.answer_callback_query(call.id, "❌ Please select at least one field!")
            return

        try:
            selected_fields = list(user_data['chart_selections'])
            chart_buffer = viewer.create_line_chart(selected_fields, code)
            if chart_buffer is None:
                bot.answer_callback_query(call.id, "❌ No data found for the specified criteria!")
                return

            # Send chart
            bot.send_photo(
                call.message.chat.id,
                chart_buffer,
                caption=f"📈 Line Chart{' for ' + code if code else ''}"
            )

            bot.answer_callback_query(call.id, "✅ Chart generated successfully!")

        except Exception as e:
            bot.answer_callback_query(call.id, f"❌ Error generating chart: {str(e)}")
            logger.error(f"Error generating chart: {e}")

        return

    # Update the message with current selections
    selected_text = ", ".join(sorted(user_data['chart_selections'])) if user_data['chart_selections'] else "None"
    updated_text = f"📊 Chart Configuration\n\nSelected fields: {selected_text}\n\nSelect fields to include in the chart:"

    bot.edit_message_text(
        updated_text,
        call.message.chat.id,
        call.message.message_id,
        reply_markup=call.message.reply_markup
    )
    bot.answer_callback_query(call.id)

if __name__ == "__main__":
    print("🤖 Bot started successfully!")
    print(f"📁 Data folder: {viewer.data_folder}")
    
    data_info = viewer.get_data_info()
    if isinstance(data_info['regular'], str):
        print("❌ No data loaded - make sure Excel files are in the 'data' folder")
    else:
        regular = data_info['regular']
        print(f"✅ Loaded {regular['total_records']} records from {regular['unique_codes']} stocks")
        print(f"📅 Date range: {regular['date_range']}")

    print("📝 Remember to replace BOT_TOKEN with your actual bot token")
    bot.infinity_polling()
    