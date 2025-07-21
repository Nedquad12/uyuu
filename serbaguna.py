import pandas as pd
from datetime import datetime
import os
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.dates as mdates
import telebot
from telebot import types
import yfinance as yf
import io
import logging
import glob
import time
import re
import gc
from collections import deque
request_queue = deque()
# ✅ Decorator untuk queue
def queued_handler(func):
    def wrapper(message, *args, **kwargs):
        user_id = message.from_user.id

        # Tambahkan user ke queue
        request_queue.append(user_id)
        queue_position = list(request_queue).index(user_id) + 1

        # Kalau bukan giliran user ini, kirim notif antrian
        if queue_position > 1:
            bot.send_message(
                message.chat.id,
                f"⏳ Anda berada di antrian ke-{queue_position}. Harap tunggu giliran..."
            )

        while request_queue[0] != user_id:
            # Tunggu giliran (cek setiap 1 detik biar tidak blocking keras)
            time.sleep(1)

        try:
            # Saat giliran, jalankan fungsi asli
            return func(message, *args, **kwargs)
        finally:
            # Hapus user dari queue setelah selesai
            if request_queue and request_queue[0] == user_id:
                request_queue.popleft()
    return wrapper

user_last_request_time = {}

def cooldown(seconds=3):
    def decorator(func):
        def wrapper(message, *args, **kwargs):
            user_id = message.from_user.id
            now = time.time()
            if user_id in user_last_request_time:
                elapsed = now - user_last_request_time[user_id]
                if elapsed < seconds:
                    bot.reply_to(message, f"⏳ Tunggu {seconds - int(elapsed)} detik sebelum request lagi.")
                    return
            user_last_request_time[user_id] = now
            return func(message, *args, **kwargs)
        return wrapper
    return decorator


WHITELIST_USER_IDS = [6208519947, 5751902978, 5209950927, 1086004279]  #ID user yang boleh DM

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize bot with your token
BOT_TOKEN ="8091313212:AAFC_d3PEMhnyOA6KzETjQ9PytAMTor8pWg"
bot = telebot.TeleBot(BOT_TOKEN)

ADMIN_LOG_FILE = "/home/nedquad12/user.xlsx"
def log_user_activity(user_id, username, command):
    """Log user activity for admin tracking"""
    try:
        # Load existing data or create new
        if os.path.exists(ADMIN_LOG_FILE):
            df = pd.read_excel(ADMIN_LOG_FILE)
        else:
            df = pd.DataFrame(columns=['user_id', 'username', 'total_requests', 'last_command', 'last_activity'])
        
        # Check if user exists
        if user_id in df['user_id'].values:
            # Update existing user
            idx = df[df['user_id'] == user_id].index[0]
            df.loc[idx, 'total_requests'] += 1
            df.loc[idx, 'last_command'] = command
            df.loc[idx, 'last_activity'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            df.loc[idx, 'username'] = username  # Update username in case it changed
        else:
            # Add new user
            new_row = pd.DataFrame({
                'user_id': [user_id],
                'username': [username],
                'total_requests': [1],
                'last_command': [command],
                'last_activity': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            })
            df = pd.concat([df, new_row], ignore_index=True)
        
        # Save to Excel
        df.to_excel(ADMIN_LOG_FILE, index=False)
        
    except Exception as e:
        logger.error(f"Error logging user activity: {e}")

class TelegramStockDataViewer:
    def __init__(self, data_folder=None):
         # Automatically set folder path
        if data_folder:
            self.data_folder = data_folder
        else:
            self.data_folder = "/home/nedquad12/database/data"  # Default relative folder
            
        self.margin_folder = "/home/nedquad12/database/margin"
        self.margin_df = None
        self.margin_fields = ['Volume', 'Nilai', 'Frekuensi']
        self.watchlist_folder = "/home/nedquad12/database/wl"
        self.chart_folder = "/home/nedquad12/database/foreign"
        
        # BlackRock folders
        self.blackrock_folders = {
            'indonesia': "/home/nedquad12/database/br/ind",
            'btc': "/home/nedquad12/database/br/btc"
        }
        
        # Data storage
        self.combined_df = None
        self.user_data = {}  # Store user-specific data (like chart selections)
        
        # BlackRock data storage
        self.blackrock_data = {
            'indonesia': None,
            'btc': None
        }
        
        self.watchlist_data = None
        self.watchlist_averages = None

        # Available fields for plotting
        self.plot_fields = [
            'Local IS', 'Local CP', 'Local PF', 'Local IB', 'Local ID', 'Local MF', 'Local SC',
            'Foreign IS', 'Foreign CP', 'Foreign PF', 'Foreign IB', 'Foreign ID', 'Foreign MF', 'Foreign SC'
        ]
        
        # Nama cantik untuk tombol Telegram
        self.button_labels = {
            'Local IS': '🇮🇩 Lokal Asuransi',
            'Local CP': '🇮🇩 Lokal Korporat',
            'Local PF': '🇮🇩 Lokal Dana Pensiun',
            'Local IB': '🇮🇩 Lokal Bank',
            'Local ID': '🇮🇩 Lokal Ritel',
            'Local MF': '🇮🇩 Lokal Reksadana',
            'Local SC': '🇮🇩 Lokal Sekuritas',
            'Foreign IS': '🌏 Asing Asuransi',
            'Foreign CP': '🌏 Asing Korporat',
            'Foreign PF': '🌏 Asing Dana Pensiun',
            'Foreign IB': '🌏 Asing Bank',
            'Foreign ID': '🌏 Asing Ritel',
            'Foreign MF': '🌏 Asing Reksadana',
            'Foreign SC': '🌏 Asing Sekuritas'
             }
    
    def load_all_excel_files(self):
        """Load all Excel files from the data folder"""
        try:
            # Create data folder if it doesn't exist
            if not os.path.exists(self.data_folder):
                os.makedirs(self.data_folder)
                logger.info(f"📂 Created data folder: {self.data_folder}")
                return

            # Find all Excel files in the data folder
            excel_files = []
            for extension in ['*.xlsx', '*.xls', '*.XLSX', '*.XLS']:
                excel_files.extend(glob.glob(os.path.join(self.data_folder, extension)))

            if not excel_files:
                logger.warning("⚠️ No Excel files found in data folder")
                return

            logger.info(f"📄 Found {len(excel_files)} Excel files: {[os.path.basename(f) for f in excel_files]}")

            # Load all Excel files
            dataframes = []
            loaded_files = []

            for file_path in excel_files:
                try:
                    logger.info(f"📥 Loading file: {file_path}")
                    df = pd.read_excel(file_path)  # FIX: Use correct file path
                    dataframes.append(df)
                    loaded_files.append(os.path.basename(file_path))
                    logger.info(f"✅ Loaded: {os.path.basename(file_path)} - {len(df)} records")
                except Exception as e:
                    logger.error(f"❌ Error loading {file_path}: {e}")
                    continue

            if dataframes:
                # Combine all dataframes
                self.combined_df = self.combine_dataframes(dataframes)
                logger.info(f"✅ Successfully loaded {len(loaded_files)} files with {len(self.combined_df)} total records")
                logger.info(f"📅 Date range: {self.combined_df['Date'].min()} to {self.combined_df['Date'].max()}")
            else:
                logger.error("❌ No valid Excel files could be loaded")
                
            self.margin_df = None  

        except Exception as e:
            logger.error(f"❌ Error during auto-load: {e}")
            
    def load_margin_files(self):
        """Load margin trading files from margin folder"""
        try:
            if not os.path.exists(self.margin_folder):
                os.makedirs(self.margin_folder)
                logger.info(f"📂 Created margin folder: {self.margin_folder}")
                return

        # Find Excel files with ddmmyy.xlsx pattern
            excel_files = []
            for extension in ['*.xlsx', '*.xls']:
                excel_files.extend(glob.glob(os.path.join(self.margin_folder, extension)))

            if not excel_files:
               logger.warning("⚠️ No margin Excel files found")
               return

        # Limit to 60 files and sort by date
            excel_files = sorted(excel_files)[:60]
            logger.info(f"📄 Found {len(excel_files)} margin files")

            dataframes = []
            for file_path in excel_files:
                try:
                # Extract date from filename (ddmmyy.xlsx)
                    filename = os.path.basename(file_path)
                    date_str = filename.split('.')[0]
                
                    df = pd.read_excel(file_path)
                
                # Add date column based on filename
                    if len(date_str) == 6:  # ddmmyy format
                        day = int(date_str[:2])
                        month = int(date_str[2:4])
                        year = int('20' + date_str[4:6])  # assume 20xx
                        file_date = datetime(year, month, day)
                        df['Date'] = file_date
                
                    dataframes.append(df)
                    logger.info(f"✅ Loaded margin file: {filename}")
                
                except Exception as e:
                    logger.error(f"❌ Error loading margin file {file_path}: {e}")
                    continue

            if dataframes:
                self.margin_df = pd.concat(dataframes, ignore_index=True)
                self.margin_df = self.margin_df.sort_values('Date', ascending=True)
                logger.info(f"✅ Loaded {len(self.margin_df)} margin records")

        except Exception as e:
            logger.error(f"❌ Error loading margin files: {e}")

    def load_blackrock_data(self):
        """Load BlackRock data from all folders"""
        for region, folder_path in self.blackrock_folders.items():
            try:
                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)
                    logger.info(f"📂 Created BlackRock folder: {folder_path}")
                    continue

                # Find Excel files
                excel_files = []
                for extension in ['*.xlsx', '*.xls']:
                    excel_files.extend(glob.glob(os.path.join(folder_path, extension)))

                if not excel_files:
                    logger.warning(f"⚠️ No BlackRock files found in {region}")
                    continue

                # Sort files by date (newest first) and limit to 60
                excel_files = sorted(excel_files, reverse=True)[:60]
                logger.info(f"📄 Found {len(excel_files)} BlackRock files for {region}")

                dataframes = []
                for file_path in excel_files:
                    try:
                        # Extract date from filename (ddmmyy.xlsx)
                        filename = os.path.basename(file_path)
                        date_str = filename.split('.')[0]
                        
                        df = pd.read_excel(file_path)
                        
                        # Add date column based on filename
                        if len(date_str) == 6:  # ddmmyy format
                            day = int(date_str[:2])
                            month = int(date_str[2:4])
                            year = int('20' + date_str[4:6])  # assume 20xx
                            file_date = datetime(year, month, day)
                            df['Date'] = file_date
                        
                        dataframes.append(df)
                        logger.info(f"✅ Loaded BlackRock file: {filename} for {region}")
                        
                    except Exception as e:
                        logger.error(f"❌ Error loading BlackRock file {file_path}: {e}")
                        continue

                if dataframes:
                    self.blackrock_data[region] = pd.concat(dataframes, ignore_index=True)
                    self.blackrock_data[region] = self.blackrock_data[region].sort_values('Date', ascending=True)
                    logger.info(f"✅ Loaded {len(self.blackrock_data[region])} BlackRock records for {region}")

            except Exception as e:
                logger.error(f"❌ Error loading BlackRock data for {region}: {e}")
    
    def load_watchlist_data(self):
        """Load watchlist data from data folder for analysis"""
        try:
            if not os.path.exists(self.watchlist_folder):
                logger.warning("⚠️ Data folder not found for watchlist")
                return

           # Find Excel files with ddmmyy.xlsx pattern
            excel_files = []
            for extension in ['*.xlsx', '*.xls']:
                excel_files.extend(glob.glob(os.path.join(self.watchlist_folder, extension)))

            if not excel_files:
                logger.warning("⚠️ No Excel files found for watchlist")
                return

        # Sort files by date (newest first) and limit to 60
            excel_files = sorted(excel_files, reverse=True)[:60]
            logger.info(f"📄 Found {len(excel_files)} files for watchlist")

            dataframes = []
            for file_path in excel_files:
                try:
                # Extract date from filename (ddmmyy.xlsx)
                    filename = os.path.basename(file_path)
                    date_str = filename.split('.')[0]
                
                    df = pd.read_excel(file_path)
                
                # Check if required columns exist
                    required_cols = ['Kode Saham', 'Penutupan', 'Volume', 'Frekuensi', 
                                   'Foreign Buy', 'Foreign Sell', 'Listed Shares']
                    missing_cols = [col for col in required_cols if col not in df.columns]
                
                    if missing_cols:
                       logger.warning(f"Missing columns in {filename}: {missing_cols}")
                       continue
                
                # Add date column based on filename
                    if len(date_str) == 6:  # ddmmyy format
                        day = int(date_str[:2])
                        month = int(date_str[2:4])
                        year = int('20' + date_str[4:6])  # assume 20xx
                        file_date = datetime(year, month, day)
                        df['Date'] = file_date
                
                # Select only required columns
                    df = df[required_cols + ['Date']]
                    dataframes.append(df)
                    logger.info(f"✅ Loaded watchlist file: {filename}")
                
                except Exception as e:
                    logger.error(f"❌ Error loading watchlist file {file_path}: {e}")
                    continue

            if dataframes:
                self.watchlist_data = pd.concat(dataframes, ignore_index=True)
                self.watchlist_data = self.watchlist_data.sort_values('Date', ascending=False)
            
            # Calculate averages
                self.calculate_watchlist_averages()
                logger.info(f"✅ Loaded {len(self.watchlist_data)} watchlist records")

        except Exception as e:
           logger.error(f"❌ Error loading watchlist data: {e}")
    
    def reload_data(self):
        """Reload all data from the data folder"""
        self.combined_df = None
        self.margin_df = None
        self.load_all_excel_files()
        self.load_margin_files()
        self.load_blackrock_data()
    
    def get_user_data(self, user_id):
        if user_id not in self.user_data:
            self.user_data[user_id] = {
                'chart_selections': set()
            }
        return self.user_data[user_id]
    
    def load_excel_file(self, file_path):
        try:
            df = pd.read_excel(file_path)
            return df
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return None
    
    def combine_dataframes(self, dfs):
        try:
        # Combine all dataframes
           combined_df = pd.concat(dfs, ignore_index=True)
        
        # Convert Date column
           combined_df['Date'] = pd.to_datetime(combined_df['Date'])
        
        # Remove duplicates based on Date and Code
           combined_df = combined_df.drop_duplicates(subset=['Date', 'Code'], keep='last')
        
        # Sort by date (oldest first - ascending order)
           combined_df = combined_df.sort_values('Date', ascending=True)
        
           return combined_df
        except Exception as e:
           logger.error(f"Error combining dataframes: {e}")
        return None
    
    def search_margin_stock(self, code):
        """Search margin data for specific stock"""
        if self.margin_df is None:
            return None
    
        stock_data = self.margin_df[self.margin_df['Kode Saham'].str.upper() == code.upper()]
        return stock_data if not stock_data.empty else None
    
    def search_stock(self, code, limit=6):
        if self.combined_df is None:
           return None
    
    # Filter data by code and limit to specified records (oldest first)
        stock_data = self.combined_df[self.combined_df['Code'].str.upper() == code.upper()].head(limit)
    
        if stock_data.empty:
           return None
    
    # Debug: Print raw data to check for duplicates
        logger.info(f"Raw data for {code}:\n{stock_data[['Date', 'Code']]}")
    
        return stock_data
    
    def search_blackrock_ticker(self, region, ticker):
        """Search BlackRock data for specific ticker"""
        if region not in self.blackrock_data or self.blackrock_data[region] is None:
            return None
        
        ticker_data = self.blackrock_data[region][
            self.blackrock_data[region]['Ticker'].str.upper() == ticker.upper()
        ]
        return ticker_data if not ticker_data.empty else None
    
    def calculate_watchlist_averages(self):
        """Calculate average Volume and Frekuensi from all data"""
        if self.watchlist_data is None:
            return
     
        try:
        # Calculate averages across all data
            self.watchlist_averages = {
                'avg_volume': self.watchlist_data['Volume'].mean(),
                'avg_frekuensi': self.watchlist_data['Frekuensi'].mean()
        }
        
            logger.info(f"📊 Calculated averages - Volume: {self.watchlist_averages['avg_volume']:,.0f}, Frekuensi: {self.watchlist_averages['avg_frekuensi']:,.0f}")
        
        except Exception as e:
            logger.error(f"❌ Error calculating averages: {e}")

    def get_watchlist_stocks(self, cap_filter=None):
        """Get stocks that meet watchlist criteria"""
        if self.watchlist_data is None or self.watchlist_averages is None:
            return []
    
        try:
        # Get latest data for each stock
            latest_data = self.watchlist_data.groupby('Kode Saham').first().reset_index()
        
        # Calculate thresholds (70% above average)
            volume_threshold = self.watchlist_averages['avg_volume'] * 1.7
            frekuensi_threshold = self.watchlist_averages['avg_frekuensi'] * 1.7
        
        # Filter stocks meeting criteria
            filtered_stocks = latest_data[
                (latest_data['Volume'] >= volume_threshold) & 
                (latest_data['Frekuensi'] >= frekuensi_threshold)
            ].copy()
        
        # Calculate additional metrics
            filtered_stocks['Net Foreign'] = filtered_stocks['Foreign Buy'] - filtered_stocks['Foreign Sell']
            filtered_stocks['Market Cap'] = filtered_stocks['Penutupan'] * filtered_stocks['Listed Shares']
        
        # Apply market cap filter
            if cap_filter:
                if cap_filter == 'high':
                    filtered_stocks = filtered_stocks[filtered_stocks['Market Cap'] >= 20e12]  # ≥20T
                elif cap_filter == 'mid':
                    filtered_stocks = filtered_stocks[
                       (filtered_stocks['Market Cap'] >= 1e12) & 
                       (filtered_stocks['Market Cap'] < 20e12)
                    ]  # ≥1T and <20T
                elif cap_filter == 'low':
                    filtered_stocks = filtered_stocks[
                        (filtered_stocks['Market Cap'] >= 80e9) & 
                        (filtered_stocks['Market Cap'] < 1e12)
                    ]  # ≥80M and <1T
                elif cap_filter == 'micro':
                    filtered_stocks = filtered_stocks[filtered_stocks['Market Cap'] < 80e9]  # <80M
                     
            foreign_60d = (
                self.watchlist_data.groupby('Kode Saham')
                .apply(lambda x: (x.sort_values('Date', ascending=False).head(60)['Foreign Buy'].sum() -
                            x.sort_values('Date', ascending=False).head(60)['Foreign Sell'].sum()))
               .reset_index(name='Net Foreign 60D')
          ) 

        # Gabungkan hasil ke filtered_stocks
            filtered_stocks = filtered_stocks.merge(foreign_60d, on='Kode Saham', how='left')
        
            return filtered_stocks.to_dict('records')
        
        except Exception as e:
            logger.error(f"❌ Error getting watchlist stocks: {e}")
            return []

    def get_foreign_flow_data(self, stock_code, days=30):
        """Get foreign flow data for specific stock"""
        if self.watchlist_data is None:
            return None
    
        try:
            stock_data = self.watchlist_data[
                self.watchlist_data['Kode Saham'].str.upper() == stock_code.upper()
            ].head(days)
        
            if stock_data.empty:
                return None
        
            stock_data = stock_data.copy()
            stock_data['Net Foreign'] = stock_data['Foreign Buy'] - stock_data['Foreign Sell']
        
            return stock_data[['Date', 'Foreign Buy', 'Foreign Sell', 'Net Foreign']].to_dict('records')
        
        except Exception as e:
            logger.error(f"❌ Error getting foreign flow data: {e}")
            return None
    
    def get_all_stock_codes(self):
        """Get all unique stock codes from the data"""
        if self.combined_df is None:
           return []

        if 'Code' not in self.combined_df.columns:
           logger.error("❌ Column 'Code' not found in data")
           return []

        try:
          # Pastikan hanya ambil nilai string
           codes = self.combined_df['Code'].dropna()
           codes = codes[codes.apply(lambda x: isinstance(x, str))]
           return sorted(codes.unique())
        except Exception as e:
           logger.error(f"❌ Error getting stock codes: {e}")
           return []
   
    def get_data_info(self):
        """Get information about the loaded data"""
        # Regular data info
        if self.combined_df is None:
            regular_info = "No regular data loaded"
        else:
            total_records = len(self.combined_df)
            unique_codes = len(self.combined_df['Code'].unique())
            date_range = f"{self.combined_df['Date'].min().strftime('%d-%b-%Y')} to {self.combined_df['Date'].max().strftime('%d-%b-%Y')}"
            regular_info = {
                'total_records': total_records,
                'unique_codes': unique_codes,
                'date_range': date_range
            }

        # Margin data info
        if self.margin_df is None:
            margin_info = "No margin data loaded"
        else:
            margin_records = len(self.margin_df)
            margin_codes = len(self.margin_df['Kode Saham'].unique())
            margin_range = f"{self.margin_df['Date'].min().strftime('%d-%b-%Y')} to {self.margin_df['Date'].max().    strftime('%d-%b-%Y')}"
            margin_info = {
                'total_records': margin_records,
                'unique_codes': margin_codes,
                'date_range': margin_range
            }

        return {
            'regular': regular_info,
            'margin': margin_info
        }   
    
    def format_stock_data(self, data):
        if data is None or data.empty:
            return "No data found"
        
        local_columns = ['Local IS', 'Local CP', 'Local PF', 'Local IB', 'Local ID', 'Local MF', 'Local SC']
        foreign_columns = ['Foreign IS', 'Foreign CP', 'Foreign PF', 'Foreign IB', 'Foreign ID', 'Foreign MF', 'Foreign SC']
        
        result = []
        for _, row in data.iterrows():
            # Calculate totals
            local_total = sum(row.get(col, 0) for col in local_columns if pd.notna(row.get(col, 0)))
            foreign_total = sum(row.get(col, 0) for col in foreign_columns if pd.notna(row.get(col, 0)))
            
            result.append(f"📅 {row['Date'].strftime('%d-%b-%Y')}")
            result.append(f"🏷️ Code: {row.get('Code', '')}")
            result.append(f"📊 Type: {row.get('Type', '')}")
            result.append(f"💰 Price: {row.get('Price', 0):,.0f}")
            result.append(f"🏠 Total Local: {local_total:,.0f}")
            result.append(f"🌍 Total Foreign: {foreign_total:,.0f}")
            result.append("─" * 30)
        
        return "\n".join(result)
    
    def create_margin_charts(self, code):
        """Create 3 separate bar charts for Volume, Nilai, Frekuensi"""
        margin_data = self.search_margin_stock(code)
        if margin_data is None:
            return None

        grouped = margin_data.groupby('Date')[self.margin_fields].sum().sort_index()

        fig, axes = plt.subplots(3, 1, figsize=(12, 15))

        # Volume Chart (Bar)
        axes[0].bar(grouped.index, grouped['Volume'], color='blue')
        axes[0].set_title(f'Volume - {code}', fontsize=14, fontweight='bold')
        axes[0].set_ylabel('Volume')
        axes[0].grid(True, alpha=0.3)

    # Nilai Chart (Bar)
        axes[1].bar(grouped.index, grouped['Nilai'], color='green')
        axes[1].set_title(f'Nilai - {code}', fontsize=14, fontweight='bold')
        axes[1].set_ylabel('Nilai')
        axes[1].grid(True, alpha=0.3)

    # Frekuensi Chart (Bar)
        axes[2].bar(grouped.index, grouped['Frekuensi'], color='red')
        axes[2].set_title(f'Frekuensi - {code}', fontsize=14, fontweight='bold')
        axes[2].set_ylabel('Frekuensi')
        axes[2].set_xlabel('Date')
        for ax in axes:
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%d%m'))
        axes[2].grid(True, alpha=0.3)

        plt.tight_layout()

      # Watermark
        plt.text(0.5, 0.5, 'Membahas Saham Indonesia', fontsize=60, color='gray',
                 ha='center', va='center', alpha=0.2, rotation=30,
                 transform=plt.gcf().transFigure, zorder=10)

        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=300, bbox_inches='tight')
        buf.seek(0)
        plt.close('all')
        gc.collect()

        return buf
    
    def create_excel_report(self, data, file_path, code):
        wb = Workbook()
        ws = wb.active
        ws.title = f"{code} Analysis"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center')
        
        # Headers
        headers = ['Date', 'Code', 'Type', 'Price', 'Local IS', 'Local CP', 'Local PF', 'Local IB', 
                  'Local ID', 'Local MF', 'Local SC', 'Foreign IS', 'Foreign CP', 'Foreign PF', 
                  'Foreign IB', 'Foreign ID', 'Foreign MF', 'Foreign SC']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        # Add Total columns headers
        ws.cell(row=1, column=len(headers)+1, value="Total Local").font = header_font
        ws.cell(row=1, column=len(headers)+1).fill = header_fill
        ws.cell(row=1, column=len(headers)+1).border = border
        ws.cell(row=1, column=len(headers)+1).alignment = center_align
        
        ws.cell(row=1, column=len(headers)+2, value="Total Foreign").font = header_font
        ws.cell(row=1, column=len(headers)+2).fill = header_fill
        ws.cell(row=1, column=len(headers)+2).border = border
        ws.cell(row=1, column=len(headers)+2).alignment = center_align
        
        # Write data
        local_cols = ['Local IS', 'Local CP', 'Local PF', 'Local IB', 'Local ID', 'Local MF', 'Local SC']
        foreign_cols = ['Foreign IS', 'Foreign CP', 'Foreign PF', 'Foreign IB', 'Foreign ID', 'Foreign MF', 'Foreign SC']
        
        for row_idx, (_, row_data) in enumerate(data.iterrows(), 2):
            # Basic data
            ws.cell(row=row_idx, column=1, value=row_data['Date'].strftime('%d-%b-%Y')).border = border
            ws.cell(row=row_idx, column=2, value=row_data.get('Code', '')).border = border
            ws.cell(row=row_idx, column=3, value=row_data.get('Type', '')).border = border
            ws.cell(row=row_idx, column=4, value=row_data.get('Price', 0)).border = border
            
            # Local and Foreign data
            for col_idx, header in enumerate(headers[4:], 5):
                value = row_data.get(header, 0)
                ws.cell(row=row_idx, column=col_idx, value=value).border = border
            
            # Add formulas for totals
            local_range = f"E{row_idx}:K{row_idx}"  # Local columns
            foreign_range = f"L{row_idx}:R{row_idx}"  # Foreign columns
            
            total_local_cell = ws.cell(row=row_idx, column=len(headers)+1)
            total_local_cell.value = f"=SUM({local_range})"
            total_local_cell.border = border
            
            total_foreign_cell = ws.cell(row=row_idx, column=len(headers)+2)
            total_foreign_cell.value = f"=SUM({foreign_range})"
            total_foreign_cell.border = border
        
        # Add summary section
        summary_row = len(data) + 3
        ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=14)
        ws.cell(row=summary_row+1, column=1, value=f"Stock Code: {code}")
        ws.cell(row=summary_row+2, column=1, value=f"Total Records: {len(data)}")
        ws.cell(row=summary_row+3, column=1, value=f"Date Range: {data['Date'].min().strftime('%d-%b-%Y')} to {data['Date'].max().strftime('%d-%b-%Y')}")
        
        # Add grand total formulas
        last_row = len(data) + 1
        ws.cell(row=summary_row+5, column=1, value="Grand Total Local:").font = Font(bold=True)
        ws.cell(row=summary_row+5, column=2, value=f"=SUM(S2:S{last_row})").font = Font(bold=True)
        ws.cell(row=summary_row+6, column=1, value="Grand Total Foreign:").font = Font(bold=True)
        ws.cell(row=summary_row+6, column=2, value=f"=SUM(T2:T{last_row})").font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    
    def create_line_chart(self, selected_fields, code=None):
        if self.combined_df is None:
            return None
        
        # Make a copy of the dataframe
        df = self.combined_df.copy()
        
        # Calculate totals if needed
        if 'Total Local' in selected_fields:
            df['Total Local'] = df[[f for f in self.plot_fields if f.startswith('Local')]].sum(axis=1)
        if 'Total Foreign' in selected_fields:
            df['Total Foreign'] = df[[f for f in self.plot_fields if f.startswith('Foreign')]].sum(axis=1)
        
        # Filter stock if applicable
        if code:
            df = df[df['Code'].str.upper() == code.upper()]
            if df.empty:
                return None
            
            
        
        # Group by date and sum
        grouped = df.groupby('Date')[selected_fields].sum().sort_index()
        
        plt.figure(figsize=(12, 8))
        for field in selected_fields:
            plt.plot(grouped.index, grouped[field], marker='o', label=field, linewidth=2)
        
        plt.xlabel('Date', fontsize=12)
        plt.ylabel('Value', fontsize=12)
        plt.title(f"Line Chart{' for ' + code if code else ''}", fontsize=14, fontweight='bold')
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        
        plt.text(0.5, 0.5, 'Membahas Saham Indonesia', fontsize=60, color='gray',
            ha='center', va='center', alpha=0.2, rotation=30,
            transform=plt.gcf().transFigure, zorder=10)
        
        # Save to bytes
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=300, bbox_inches='tight')
        buf.seek(0)
        plt.close('all')
        gc.collect()
        
        return buf
    
    def create_blackrock_chart(self, region, ticker):
        """Create chart for BlackRock ticker data"""
        ticker_data = self.search_blackrock_ticker(region, ticker)
        if ticker_data is None:
            return None, None
        
        # Group by date and get latest values for each date
        grouped = ticker_data.groupby('Date').last().sort_index()
        
        # Create chart for Quantity Total
        plt.figure(figsize=(12, 8))
        plt.plot(grouped.index, grouped['Quantity Total'], marker='o', linewidth=2, color='blue')
        plt.title(f'BlackRock Holdings - {ticker} ({region.upper()})', fontsize=14, fontweight='bold')
        plt.xlabel('Date', fontsize=12)
        plt.ylabel('Quantity Total', fontsize=12)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        
        plt.text(0.5, 0.5, 'Membahas Saham Indonesia', fontsize=60, color='gray',
         ha='center', va='center', alpha=0.2, rotation=30,
         transform=plt.gcf().transFigure, zorder=10)
        
        # Save chart to buffer
        chart_buf = io.BytesIO()
        plt.savefig(chart_buf, format='png', dpi=300, bbox_inches='tight')
        chart_buf.seek(0)
        plt.close('all')
        gc.collect()
        
        # Generate movement caption
        caption = self.generate_movement_caption(grouped, ticker, region)
        
        return chart_buf, caption

    def generate_movement_caption(self, grouped_data, ticker, region):
        """Generate detailed movement caption"""
        if len(grouped_data) < 2:
            return f"📊 BlackRock Holdings - {ticker} ({region.upper()})\n❌ Insufficient data for movement analysis"
        
        # Get latest and previous data
        latest = grouped_data.iloc[-1]
        previous = grouped_data.iloc[-2]
        
        # Calculate changes
        qty_change = latest['Quantity Total'] - previous['Quantity Total']
        qty_change_pct = (qty_change / previous['Quantity Total']) * 100 if previous['Quantity Total'] != 0 else 0
        
        mv_change = latest['Market Value Total'] - previous['Market Value Total']
        mv_change_pct = (mv_change / previous['Market Value Total']) * 100 if previous['Market Value Total'] != 0 else 0
        
        # Format numbers
        qty_latest = f"{latest['Quantity Total']:,.0f}"
        qty_prev = f"{previous['Quantity Total']:,.0f}"
        mv_latest = f"{latest['Market Value Total']:,.0f}"
        mv_prev = f"{previous['Market Value Total']:,.0f}"
        
        # Direction indicators
        qty_arrow = "🔺" if qty_change > 0 else "🔻" if qty_change < 0 else "➡️"
        mv_arrow = "🔺" if mv_change > 0 else "🔻" if mv_change < 0 else "➡️"
        
        caption = f"""📊 BlackRock Holdings - {ticker} ({region.upper()})

📅 Latest: {latest.name.strftime('%d-%b-%Y')}
📅 Previous: {previous.name.strftime('%d-%b-%Y')}

📈 Quantity Total:
Current: {qty_latest}
Previous: {qty_prev}
Change: {qty_arrow} {qty_change:+,.0f} ({qty_change_pct:+.2f}%)

💰 Market Value Total:
Current: ${mv_latest}
Previous: ${mv_prev}
Change: {mv_arrow} ${mv_change:+,.0f} ({mv_change_pct:+.2f}%)"""
        
        return caption

    def get_significant_movements(self, threshold=3.0):
        """Get all tickers with significant movements (>= threshold%)"""
        movements = []
        
        for region, data in self.blackrock_data.items():
            if data is None or len(data) < 2:
                # Limit to latest 5 dates for /b7 command only
                latest_dates = sorted(data['Date'].unique(), reverse=True)[:5]
                data = data[data['Date'].isin(latest_dates)]
                continue
                
            # Get unique tickers
            tickers = data['Ticker'].unique()
            
            for ticker in tickers:
                ticker_data = data[data['Ticker'] == ticker].groupby('Date').last().sort_index()
                
                if len(ticker_data) < 2:
                    continue
                    
                # Calculate movement
                latest = ticker_data.iloc[-1]
                previous = ticker_data.iloc[-2]
                
                if previous['Quantity Total'] == 0:
                    continue
                    
                qty_change_pct = ((latest['Quantity Total'] - previous['Quantity Total']) / previous['Quantity Total']) * 100
                
                if abs(qty_change_pct) >= threshold:
                    movements.append({
                        'region': region,
                        'ticker': ticker,
                        'change_pct': qty_change_pct,
                        'latest_qty': latest['Quantity Total'],
                        'previous_qty': previous['Quantity Total'],
                        'latest_mv': latest['Market Value Total'],
                        'previous_mv': previous['Market Value Total'],
                        'latest_date': latest.name,
                        'previous_date': previous.name
                    })
        
        # Sort by absolute change percentage
        movements.sort(key=lambda x: abs(x['change_pct']), reverse=True)
        return movements

    def get_foreign_flow_data(self, stock_code, days=30):
        """Get foreign flow data for specific stock"""
        if self.watchlist_data is None:
            return None
    
        try:
            stock_data = self.watchlist_data[
                self.watchlist_data['Kode Saham'].str.upper() == stock_code.upper()
            ].head(days)
        
            if stock_data.empty:
                return None
        
            stock_data = stock_data.copy()
            stock_data['Net Foreign'] = stock_data['Foreign Buy'] - stock_data['Foreign Sell']
        
            return stock_data[['Date', 'Foreign Buy', 'Foreign Sell', 'Net Foreign']].to_dict('records')
        
        except Exception as e:
           logger.error(f"❌ Error getting foreign flow data: {e}")
           return None

    def format_watchlist_response(self, stocks, cap_filter=None):
        """Format watchlist response for Telegram"""
        if not stocks:
            return "❌ No stocks found meeting the criteria"
    
        cap_names = {
            'high': 'High Cap (≥20T)',
            'mid': 'Mid Cap (≥1T)',
            'low': 'Low Cap (≥80M)',
            'micro': 'Micro Cap (<80M)'
       }
    
        response = f"📊 WATCHLIST STOCKS"
        if cap_filter:
            response += f" - {cap_names.get(cap_filter, cap_filter.upper())}"
    
        response += f"\n\n🔍 Saham Watchlist\nNote: Bukan ajakan jual dan beli\n"
        response += f"📈 Avg Volume: {self.watchlist_averages['avg_volume']:,.0f}\n"
        response += f"📈 Avg Frekuensi: {self.watchlist_averages['avg_frekuensi']:,.0f}\n"
        response += f"📊 Threshold: Vol≥{self.watchlist_averages['avg_volume']*1.7:,.0f}, Freq≥{self.watchlist_averages['avg_frekuensi']*1.7:,.0f}\n\n"
    
        # Sort by market cap descending
        stocks_sorted = sorted(stocks, key=lambda x: x['Market Cap'], reverse=True)
     
        for i, stock in enumerate(stocks_sorted, 1): 
        # Format market cap
            market_cap = stock['Market Cap']
            if market_cap >= 1e12:
                cap_str = f"{market_cap/1e12:.1f}T"
            elif market_cap >= 1e9:
                cap_str = f"{market_cap/1e9:.1f}B"
            else:
                cap_str = f"{market_cap/1e6:.1f}M"
        
        # Format net foreign
            net_foreign_today = stock.get('Net Foreign', 0)
            foreign_arrow = "🔺" if net_foreign_today > 0 else "🔻" if net_foreign_today < 0 else "➡️"

            net_foreign_60d = stock.get('Net Foreign 60D', 0)
            foreign_60d_arrow = "🔺" if net_foreign_60d > 0 else "🔻" if net_foreign_60d < 0 else "➡️"
        
            response += f"{i}. {stock['Kode Saham']}\n"
            response += f"   💰 Price: {stock['Penutupan']:,.0f}\n"
            response += f"   📊 Cap: {cap_str}\n"
            response += f"   📈 Vol: {stock['Volume']:,.0f}\n"
            response += f"   🔄 Freq: {stock['Frekuensi']:,.0f}\n"
            response += f"   {foreign_arrow} Net Foreign: {net_foreign_today:+,.0f}\n"
            response += f"   {foreign_60d_arrow} Net Foreign 60D: {net_foreign_60d:+,.0f}\n\n"
    
        return response
    
    def load_blackrock_data_for_region(self, region):
        """Load BlackRock data for a specific region only"""
        folder_path = self.blackrock_folders.get(region)
        if not folder_path or not os.path.exists(folder_path):
            logger.warning(f"⚠️ Folder not found for BlackRock region: {region}")
            if self.blackrock_data is None:
               self.blackrock_data = {
               'indonesia': None,
               'btc': None
          }

        excel_files = []
        for extension in ['*.xlsx', '*.xls']:
            excel_files.extend(glob.glob(os.path.join(folder_path, extension)))

        if not excel_files:
            logger.warning(f"⚠️ No BlackRock files found in {region}")
            self.blackrock_data[region] = None
            return

        excel_files = sorted(excel_files, reverse=True)[:60]
        logger.info(f"📄 Found {len(excel_files)} BlackRock files for {region}")

        dataframes = []
        for file_path in excel_files:
            try:
                df = pd.read_excel(file_path)
                filename = os.path.basename(file_path)
                date_str = filename.split('.')[0]
                if len(date_str) == 6:  # ddmmyy format
                    day = int(date_str[:2])
                    month = int(date_str[2:4])
                    year = int('20' + date_str[4:6])  # assume 20xx
                    file_date = datetime(year, month, day)
                    df['Date'] = file_date
                dataframes.append(df)
            except Exception as e:
                logger.error(f"❌ Error loading BlackRock file {file_path}: {e}")
                continue

        if dataframes:
            self.blackrock_data[region] = pd.concat(dataframes, ignore_index=True)
            self.blackrock_data[region] = self.blackrock_data[region].sort_values('Date', ascending=True)
            logger.info(f"✅ Loaded {len(self.blackrock_data[region])} BlackRock records for {region}")
        else:
            self.blackrock_data[region] = None


# Initialize the viewer
viewer = TelegramStockDataViewer()

def whitelist_only(func):
    def wrapper(message):
        if message.chat.type == 'private' and message.from_user.id not in WHITELIST_USER_IDS:
            bot.send_message(
                message.chat.id,
                "Hi!! Maaf kamu tidak terdaftar sebagai member premium. Jika ingin  bergabung menjadi member premium dan menggunakan bot sesuka hati bisa hubungi @Rendanggedang"
            )
            return
        return func(message)
    return wrapper

@bot.message_handler(commands=['start', 'help'])
@whitelist_only
def send_welcome(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    
    if isinstance(data_info['regular'], str):
        data_status = "❌ No data loaded"
    else:
        regular = data_info['regular']
        data_status = f"✅ Data Holding loaded: {regular['total_records']} records, {regular['unique_codes']} stocks\n📅 Date range: {regular['date_range']}"
    
    # Perbaikan: Ubah ke format yang lebih sederhana tanpa parsing entities
    help_text = f"""🤖 Bot Analisa Saham Indonesia, US, dan BTC.

Perintah untuk IHSG
/help - Memanggil pesan ini
/search [CODE] - Search untuk data holding saham
/export [CODE] - Export data ke Excel
/chart - Membuat 
/g - Pergerakan saham harian
/wl - Watchlist saham
/m [CODE] - Menampikan data transaksi Margin

Perintah untuk Blacrock 
/bi for Indonesia stock
/btc for Bitcoin
/b7 - Pergerakan besar

/i for index

Contoh:
- /search BBCA - Mencari data saham BBCA
- /bi BBCA - Mencari saham BBCA yang dimiliki Blackrock

Twitter Owner: https://x.com/saberial_link/
Telegram Owner: @Rendanggedang"""
    
    # Kirim tanpa parse_mode untuk menghindari error parsing
    bot.reply_to(message, help_text)

@bot.message_handler(commands=['reload'])
@whitelist_only
def reload_data(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "reload")

    try:
        bot.reply_to(message, "🔄 Bersih-bersih RAM & memuat ulang data...")

        # ✅ Bersihkan semua data di RAM
        viewer.combined_df = None
        viewer.margin_df = None
        viewer.watchlist_data = None
        viewer.watchlist_averages = None
        viewer.blackrock_data = {
            'indonesia': None,
            'btc': None
        }
        viewer.user_data = {}  # Bersihkan cache user chart selections

        import gc
        gc.collect()  # 🧹 Paksa garbage collection

        bot.reply_to(message, "✅ RAM sudah bersih.")

    except Exception as e:
        bot.reply_to(message, f"❌ Error saat reload: {str(e)}")
        logger.error(f"Error di /reload: {e}")

    
@bot.message_handler(commands=['m'])
@whitelist_only
@queued_handler
@cooldown(seconds=5)
def margin_trading(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    viewer.load_margin_files()
   
    if viewer.margin_df is None:
        bot.reply_to(message, "❌ No margin data loaded. Saham ini tidak terdaftar dalam margin.")
        return
    
    parts = message.text.split()
    if len(parts) < 2:
        bot.reply_to(message, "❌ Maskan kode saham.\nContoh: `/m BBCA`", parse_mode='Markdown')
        return
    
    code = parts[1].upper()
    
    try:
        chart_buffer = viewer.create_margin_charts(code)
        if chart_buffer is None:
            bot.reply_to(message, f"❌ Saham ini tidak termasuk daftar Margin: {code}")
            return
        
        # Send chart
        bot.send_photo(
            message.chat.id,
            chart_buffer,
            caption=f"📊 Transaction Margin for {code}\n📈 Volume, Nilai, Frekuensi"
        )
        viewer.margin_df = None
        plt.close('all')
        gc.collect()
        
    except Exception as e:
        bot.reply_to(message, f"❌ Error creating margin chart: {str(e)}")
        logger.error(f"Error creating margin chart: {e}")
    
    finally:    
       if message.from_user.id in request_queue:
          request_queue.remove(message.from_user.id)


@bot.message_handler(commands=['export'])
@whitelist_only
@queued_handler
@cooldown(seconds=5)
def export_stock(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    
    viewer.load_all_excel_files()
  
    if viewer.combined_df is None:
        bot.reply_to(message, "❌ Tidak ada data. Beritahu admin @Rendanggedang atau https://x.com/saberial_link/ bahwa server bermaslah")
        return
    
    # Extract stock code from command
    parts = message.text.split()
    if len(parts) < 2:
        bot.reply_to(message, "❌ Masukan kode saham\nContoh: /export BBCA")
        return
    
    code = parts[1].upper()
    
    # Search for stock
    stock_data = viewer.search_stock(code)
    
    if stock_data is None:
        bot.reply_to(message, f"❌ Tidak ada dat untuk: {code}")
        return
    
    try:
        # Create Excel report
        export_file = f"export_{code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        viewer.create_excel_report(stock_data, export_file, code)
        
        # Send file
        with open(export_file, 'rb') as f:
            bot.send_document(message.chat.id, f, caption=f"📊 Stock analysis for {code}")
        
        # Clean up
        os.remove(export_file)
        viewer.combined_df = None        
        
    except Exception as e:
        bot.reply_to(message, f"❌ Error creating Excel report: {str(e)}")
        logger.error(f"Error creating Excel report: {e}")
        
    finally:    
       if message.from_user.id in request_queue:
          request_queue.remove(message.from_user.id)
        
        
@bot.message_handler(commands=['g'])
@whitelist_only
def top_gainers_losers(message):
    try:
        folder_path = viewer.chart_folder  # 🆕 folder sumber data
        excel_files = sorted(glob.glob(os.path.join(folder_path, "*.xlsx")), reverse=True)
        if not excel_files:
            bot.reply_to(message, f"❌ Tidak ada file Excel di folder `{folder_path}`.")
            return

        latest_file = excel_files[0]
        df = pd.read_excel(latest_file)

        # Pastikan kolom ada
        required_cols = ['Kode Saham', 'Penutupan', 'Sebelumnya']
        if not all(col in df.columns for col in required_cols):
            bot.reply_to(message, f"❌ Kolom {required_cols} tidak ditemukan di file Excel.")
            return

        # Hitung perubahan harga dan persentase
        df['Change'] = df['Penutupan'] - df['Sebelumnya']
        df['Change %'] = (df['Change'] / df['Sebelumnya']) * 100

        # Top 20 Gainers
        top_gainers = df.sort_values('Change %', ascending=False).head(20)
        # Top 20 Losers
        top_losers = df.sort_values('Change %', ascending=True).head(20)

        # Format angka
        def format_pct(val):
            return f"{val:+.1f}%"

        top_gainers['Change %'] = top_gainers['Change %'].apply(format_pct)
        top_losers['Change %'] = top_losers['Change %'].apply(format_pct)

        # Format ke Text Table
        gainers_table = "\n📈 *Top 20 Gainers*\n\n"
        gainers_table += "Kode      Close     %Change\n"
        gainers_table += "\n".join([
            f"{row['Kode Saham']:<10}{row['Penutupan']:>8,.0f}   {row['Change %']:>6}"
            for _, row in top_gainers.iterrows()
        ])

        losers_table = "\n\n📉 *Top 20 Losers*\n\n"
        losers_table += "Kode      Close     %Change\n"
        losers_table += "\n".join([
            f"{row['Kode Saham']:<10}{row['Penutupan']:>8,.0f}   {row['Change %']:>6}"
            for _, row in top_losers.iterrows()
        ])

        # Kirim ke Telegram
        bot.reply_to(message, f"{gainers_table}{losers_table}", parse_mode='Markdown')

    except Exception as e:
        bot.reply_to(message, f"❌ Error: {str(e)}")
        logger.error(f"Error in /g command: {e}")

@bot.message_handler(commands=['bi'])
@whitelist_only
@queued_handler
@cooldown(seconds=5)
def blackrock_indonesia(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    handle_blackrock_command(message, 'indonesia')

@bot.message_handler(commands=['btc'])
@whitelist_only
@queued_handler
@cooldown(seconds=5)
def blackrock_btc(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    viewer.load_blackrock_data_for_region('btc')
    
    # Bitcoin only - tidak perlu ticker
    if viewer.blackrock_data['btc'] is None:
        bot.reply_to(message, "❌ Tidak ada data Bitcoin di server, beritahu admin @Rendanggedang atau https://x.com/saberial_link/.")
        return
    
    try:
        # Assume Bitcoin ticker for BTC
        chart_buffer, caption = viewer.create_blackrock_chart('btc', 'BTC')
        if chart_buffer is None:
            bot.reply_to(message, "❌ Tidak ada data Bitcoin di server, beritahu admin @Rendanggedang atau https://x.com/saberial_link/")
            return
        
        bot.send_photo(message.chat.id, chart_buffer, caption=caption)
        viewer.blackrock_data['btc'] = None
        
    except Exception as e:
        bot.reply_to(message, f"❌ Gagal membuat chart, server overload: {str(e)}")
        logger.error(f"Error creating BlackRock Bitcoin chart: {e}")
        
    finally:    
       if message.from_user.id in request_queue:
          request_queue.remove(message.from_user.id)

def handle_blackrock_command(message, region):
    """Handle BlackRock commands that require ticker input"""
    viewer.load_blackrock_data_for_region(region)
       
    if viewer.blackrock_data[region] is None:
        bot.reply_to(message, f"❌ No BlackRock {region} data loaded.")
        return
    
    parts = message.text.split()
    if len(parts) < 2:
        bot.reply_to(message, f"❌ Masukan kode saham.\nExamplContoh: /{parts[0][1:]} BBCA")
        return
    
    ticker = parts[1].upper()
    
    try:
        chart_buffer, caption = viewer.create_blackrock_chart(region, ticker)
        if chart_buffer is None:
            bot.reply_to(message, f"❌ Data tidak tersedia: {ticker} in {region}")
            return
        
        bot.send_photo(message.chat.id, chart_buffer, caption=caption)
        
        viewer.blackrock_data[region] = None
        
    except Exception as e:
        bot.reply_to(message, f"❌ Error creating BlackRock {region} chart: {str(e)}")
        logger.error(f"Error creating BlackRock {region} chart: {e}")
        
    finally:    
       if message.from_user.id in request_queue:
          request_queue.remove(message.from_user.id)

@bot.message_handler(commands=['b7'])
@whitelist_only
@queued_handler
@cooldown(seconds=5)
def blackrock_significant_movements(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    viewer.load_blackrock_data()
    
    """Show significant BlackRock movements (>= 3%)"""
    try:
        for region in viewer.blackrock_folders.keys():
            viewer.load_blackrock_data_for_region(region)
        movements = viewer.get_significant_movements(3.0)
        
        if not movements:
            bot.reply_to(message, "📊 No significant BlackRock movements (>= 3%) found in the last period.")
            return
        
        # Format response
        response = "📊 Pergerakan Signifikan Blackrock\n\n"
        
        for i, movement in enumerate(movements[:20]):  # Limit to top 20
            arrow = "🔺" if movement['change_pct'] > 0 else "🔻"
            response += f"{i+1}. {movement['ticker']} ({movement['region'].upper()})\n"
            response += f"   {arrow} {movement['change_pct']:+.2f}%\n"
            response += f"   Latest: {movement['latest_date'].strftime('%d-%b-%Y')}\n"
            response += f"   Previous: {movement['previous_date'].strftime('%d-%b-%Y')}\n"
            response += f"   Qty: {movement['latest_qty']:,.0f}\n"
            response += f"   MV: ${movement['latest_mv']:,.0f}\n\n"
        
        if len(movements) > 20:
            response += f"... and {len(movements) - 20} more movements"
        
        bot.reply_to(message, response)
        viewer.blackrock_data = {
        'indonesia': None,
        'btc': None
         }

        for region in viewer.blackrock_folders.keys():
            viewer.blackrock_data[region] = None
        
    except Exception as e:
        bot.reply_to(message, f"❌ Error getting significant movements: {str(e)}")
        logger.error(f"Error getting significant movements: {e}")
        
@bot.message_handler(commands=['search'])
@whitelist_only
def search_stock_command(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    
    viewer.load_all_excel_files()
    
    parts = message.text.split()
    if len(parts) < 2:
        bot.reply_to(message, "❌ Masukan kode saham.\nContoh: `/search BBCA`", parse_mode='Markdown')
        return

    code = parts[1].upper()
    stock_data = viewer.search_stock(code)

    if stock_data is None or stock_data.empty:
        bot.reply_to(message, f"❌ Tidak ada data untuk saham {code}.")
        return

    response_text = viewer.format_stock_data(stock_data)
    bot.reply_to(message, response_text)
    
    viewer.combined_df = None

@bot.message_handler(commands=['wl'])
@whitelist_only
def watchlist_command(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "watchlist")
    viewer.load_watchlist_data()
    if viewer.watchlist_data is None:
        bot.reply_to(message, "❌ Tidak ada data. API bermasalah, segera beritahu admin @Rendanggedang atau https://x.com/saberial_link/.")
        return
    
    # Create keyboard for market cap filter
    markup = types.InlineKeyboardMarkup(row_width=2)
    markup.add(
        types.InlineKeyboardButton("🔥 High Cap (≥20T)", callback_data="wl_high"),
        types.InlineKeyboardButton("📈 Mid Cap (≥1T)", callback_data="wl_mid")
    )
    markup.add(
        types.InlineKeyboardButton("📊 Low Cap (≥80M)", callback_data="wl_low"),
        types.InlineKeyboardButton("🔍 Micro Cap (<80M)", callback_data="wl_micro")
    )
    markup.add(types.InlineKeyboardButton("📋 All Caps", callback_data="wl_all"))
    
    bot.reply_to(message, "📊 Pilih Market Cap:", reply_markup=markup)
    
import yfinance as yf

@bot.message_handler(commands=['i'])
@whitelist_only
@queued_handler
@cooldown(seconds=5)
def show_indices_data(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "/i")

    try:
        indices = {
            'ES=F': 'USA',
            'BTC-USD': 'Bitcoin',
            '^NZ50': 'N.Zealand',
            '^AXJO': 'Australia',
            '^N225': 'Jepang',
            '^KS11': 'Korea',
            '^STI': 'Singapore',
            '000001.SS': 'China',
            '^HSI': 'Hong Kong',
            '^KLSE': 'Malaysia',
            '^TWII': 'Taiwan',
            '^JKSE': 'Indonesia',
            '^BSESN': 'India'
        }

        currencies = {
            'IDR=X': 'USD/IDR',
            'JPY=X': 'USD/JPY',
            'AUDUSD=X': 'AUD/USD',
            'EURUSD=X': 'EUR/USD',
            'GBPUSD=X': 'GBP/USD',
            'THB=X': 'USD/THB',
            'MYR=X': 'USD/MYR'
        }

        commodities = {
            'GC=F': 'Gold',
            'SI=F': 'Silver',
            'PL=F': 'Platinum',
            'CL=F': 'Crude Oil',
            'HG=F': 'Copper',
            'NG=F': 'Natural Gas'
        }

        response = "📊 *Data Index, Currency and Commodities.*\n\n"

        def fetch_data(tickers, category_name):
            lines = [f"📌 *{category_name}*"]
            for ticker, name in tickers.items():
                try:
                    data = yf.Ticker(ticker)
                    price = data.history(period="1d", interval="1m")['Close'].iloc[-1]
                    prev_close = data.history(period="2d")['Close'].iloc[-2]
                    change = price - prev_close
                    pct_change = (change / prev_close) * 100

                    if change > 0:
                        arrow = "🟢"
                    elif change < 0:
                        arrow = "🔴"
                    else:
                        arrow = "⚪"

                    lines.append(f"{name}: {price:,.2f} {arrow} {change:+.2f} ({pct_change:+.2f}%)")
                except Exception as e:
                    lines.append(f"{name}: ❌ Error")
            return "\n".join(lines)

        response += fetch_data(indices, "Indices")
        response += "\n\n" + fetch_data(currencies, "Currencies")
        response += "\n\n" + fetch_data(commodities, "Commodities")

        bot.reply_to(message, response, parse_mode="Markdown")

    except Exception as e:
        bot.reply_to(message, f"❌ Error fetching market data: {str(e)}")


@bot.message_handler(commands=['chart'])
@whitelist_only
@queued_handler
@cooldown(seconds=5)
def create_chart(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "search")
    viewer.load_all_excel_files()
    request_queue.append(message.from_user.id)
    queue_position = list(request_queue).index(message.from_user.id) + 1

    if queue_position > 1:
       bot.send_message(
           message.chat.id,
           f"⏳ Anda berada di antrian ke [{queue_position}] harap bersabar..."
       )
    
    if viewer.combined_df is None:
        bot.reply_to(message, "❌ Tidak ada data. API bermasalah, segera beritahu admin @Rendanggedang atau https://x.com/saberial_link/.")
        return
    
    # Extract stock code from command (optional)
    parts = message.text.split()
    code = parts[1].upper() if len(parts) > 1 else None
    
    # Create keyboard for field selection
    markup = types.InlineKeyboardMarkup(row_width=2)
    
    # Add field selection buttons
    field_buttons = []
    for field in viewer.plot_fields:
        label = viewer.button_labels.get(field, field)  # Pakai nama cantik
        callback_data = f"field_{field.replace(' ', '_')}"
        field_buttons.append(types.InlineKeyboardButton(label, callback_data=callback_data))
    
    # Add buttons in pairs
    for i in range(0, len(field_buttons), 2):
        if i + 1 < len(field_buttons):
            markup.add(field_buttons[i], field_buttons[i + 1])
        else:
            markup.add(field_buttons[i])
    
    # Add total buttons
    markup.add(
        types.InlineKeyboardButton("Total Local", callback_data="field_Total_Local"),
        types.InlineKeyboardButton("Total Foreign", callback_data="field_Total_Foreign")
    )
    
    # Add control buttons
    markup.add(
        types.InlineKeyboardButton("Select All", callback_data="select_all"),
        types.InlineKeyboardButton("Clear All", callback_data="clear_all")
    )
    markup.add(types.InlineKeyboardButton("Generate Chart", callback_data=f"generate_chart_{code if code else 'all'}"))
    
    chart_text = f"📊 Pilih data yang diinginlkan\nJika sudah jangan lupa klik Generate Chart\n"
    if code:
        chart_text += f"Stock: {code}\n"
    chart_text += "\nSelect fields to include in the chart:"
    
    # Kirim tanpa parse_mode
    bot.reply_to(message, chart_text, reply_markup=markup)
 
    request_queue.append(message.from_user.id)
    queue_position = list(request_queue).index(message.from_user.id) + 1

    if queue_position > 1:
       bot.send_message(
           message.chat.id,
           f"⏳ Anda berada di antrian ke [{queue_position}] harap bersabar..."
       )

@bot.callback_query_handler(func=lambda call: call.data.startswith('field_') or call.data in ['select_all', 'clear_all'] or call.data.startswith('generate_chart_'))

@bot.callback_query_handler(func=lambda call: call.data.startswith('field_') or call.data in ['select_all', 'clear_all'] or call.data.startswith('generate_chart_'))
def handle_chart_selection(call):
    user_id = call.from_user.id
    user_data = viewer.get_user_data(user_id)

    if call.data.startswith('field_'):
        field = call.data[6:].replace('_', ' ')
        if field in user_data['chart_selections']:
            user_data['chart_selections'].remove(field)
        else:
            user_data['chart_selections'].add(field)

    elif call.data == 'select_all':
        user_data['chart_selections'] = set(viewer.plot_fields + ['Total Local', 'Total Foreign'])

    elif call.data == 'clear_all':
        user_data['chart_selections'] = set()

    elif call.data.startswith('generate_chart_'):
        code = call.data[15:] if call.data[15:] != 'all' else None

        if not user_data['chart_selections']:
            bot.answer_callback_query(call.id, "❌ Please select at least one field!")
            return

        try:
            selected_fields = list(user_data['chart_selections'])
            chart_buffer = viewer.create_line_chart(selected_fields, code)
            if chart_buffer is None:
                bot.answer_callback_query(call.id, "❌ No data found for the specified criteria!")
                return

            # Send chart
            bot.send_photo(
                call.message.chat.id,
                chart_buffer,
                caption=f"📈 Line Chart{' for ' + code if code else ''}"
            )
            
            viewer.combined_df = None
            
            chart_buffer.close() 
            plt.close('all')      
            gc.collect()

            bot.answer_callback_query(call.id, "✅ Chart generated successfully!")

        except Exception as e:
            bot.answer_callback_query(call.id, f"❌ Error generating chart: {str(e)}")
            logger.error(f"Error generating chart: {e}")

        return

    # Update the message with current selections
    selected_text = ", ".join(sorted(user_data['chart_selections'])) if user_data['chart_selections'] else "None"
    updated_text = f"📊 Chart Configuration\n\nSelected fields: {selected_text}\n\nSelect fields to include in the chart:"

    bot.edit_message_text(
        updated_text,
        call.message.chat.id,
        call.message.message_id,
        reply_markup=call.message.reply_markup
    )
    bot.answer_callback_query(call.id)
    
@bot.message_handler(commands=['ff'])
@whitelist_only
@queued_handler
@cooldown(seconds=5)
def free_float_summary(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "ff")
    try:
        folder_path = "/home/nedquad12/database/foreign"
        excel_files = sorted(glob.glob(os.path.join(folder_path, "*.xlsx")), reverse=True)
        if not excel_files:
            bot.reply_to(message, f"❌ Tidak ada file Excel di folder `{folder_path}`.")
            return

        latest_file = excel_files[0]
        df = pd.read_excel(latest_file)

        required_cols = ['Kode Saham', 'Weight For Index', 'Penutupan', 'Listed Shares']
        if not all(col in df.columns for col in required_cols):
            bot.reply_to(message, f"❌ Kolom {required_cols} tidak ditemukan di file Excel.")
            return

        parts = message.text.split()
        if len(parts) < 2:
            bot.reply_to(message, "❌ Masukan kode saham.\nContoh: `/ff BBCA`", parse_mode='Markdown')
            return

        code = parts[1].upper()
        stock = df[df['Kode Saham'].str.upper() == code]

        if stock.empty:
            bot.reply_to(message, f"❌ Saham {code} tidak ditemukan.")
            return

        weight_index = stock['Weight For Index'].values[0]
        penutupan = stock['Penutupan'].values[0]
        listed_shares = stock['Listed Shares'].values[0]

        # Hitung Free Float Value
        ff_value = weight_index * penutupan
        ff_percent = (weight_index / listed_shares) * 100

        # Format nilai
        def format_value(val):
            if val >= 1e12:
                return f"{val/1e12:.0f}T"
            elif val >= 1e9:
                return f"{val/1e9:.1f}B"
            elif val >= 1e6:
                return f"{val/1e6:.1f}M"
            else:
                return f"{val:.0f}"

        ff_value_str = format_value(ff_value)

        # Bandingkan dengan MSCI
        kurs = 16300
        min_value_15 = 6000e9  # Rp5500B
        min_value_low_ff = 8400e12  # Rp8400T

        meets_value = ff_value * kurs >= min_value_15
        meets_ff = ff_percent >= 15

        if meets_ff and meets_value:
            status = "✅ Memenuhi syarat MSCI (Rp5500B, FF≥15%)"
        elif not meets_ff and ff_value * kurs >= min_value_low_ff:
            status = "⚠️ Potensi eligible."
        else:
            status = "❌ Belum memenuhi syarat MSCI"

        response = (
            f"📊 Free Float Summary for {code}\n"
            f"💰 FF Value: {ff_value_str}\n"
            f"📈 Free Float: {ff_percent:.2f}%\n"
            f""
        )
        bot.reply_to(message, response)

    except Exception as e:
        bot.reply_to(message, f"❌ Error: {str(e)}")
        logger.error(f"Error in /ff command: {e}")

@bot.message_handler(commands=['hol'])
@whitelist_only
@queued_handler
@cooldown(seconds=5)
def holdings_summary_with_stock(message):
    log_user_activity(message.from_user.id, message.from_user.username or "Unknown", "/hol")
    try:
        # Load data kepemilikan
        viewer.load_all_excel_files()
        if viewer.combined_df is None:
            bot.reply_to(message, "❌ Tidak ada data kepemilikan. Jalankan /reload dulu.")
            return

        parts = message.text.split()
        if len(parts) < 2:
            bot.reply_to(message, "❌ Masukan kode saham.\nContoh: `/hol BBCA`", parse_mode='Markdown')
            return

        code = parts[1].upper()
        df = viewer.search_stock(code)
        if df is None or df.empty:
            bot.reply_to(message, f"❌ Tidak ada data untuk saham {code}.")
            return

        # Load harga penutupan dari folder foreign
        folder_path = "/home/nedquad12/database/foreign"
        excel_files = sorted(glob.glob(os.path.join(folder_path, "*.xlsx")), reverse=True)
        if not excel_files:
            bot.reply_to(message, "❌ Tidak ada file harga penutupan di folder foreign.")
            return

        harga_data = pd.read_excel(excel_files[0])
        harga_dict = dict(zip(harga_data['Kode Saham'].str.upper(), harga_data['Penutupan']))

        closing_price = harga_dict.get(code)
        if closing_price is None:
            bot.reply_to(message, f"❌ Tidak ditemukan harga penutupan untuk {code}.")
            return

        df['Month'] = df['Date'].dt.to_period('M')

        # Kategori yang diminta
        ritel_columns = ['Local ID', 'Foreign ID']
        bandar_lokal_columns = ['Local IS', 'Local MF', 'Local SC', 'Local OT']
        bandar_asing_columns = ['Foreign IS', 'Foreign MF', 'Foreign SC', 'Foreign OT']
        big_investor_lokal_columns = ['Local CP', 'Local PF', 'Local IB', 'Local FD']
        big_investor_asing_columns = ['Foreign CP', 'Foreign PF', 'Foreign IB', 'Foreign FD']

        response = f"📊 Holding Summary for {code}\n💰 Harga Penutupan: Rp {closing_price:,.0f}\n"

        for month, group in df.groupby('Month'):
            ritel_total = group[ritel_columns].sum().sum() * closing_price
            bandar_lokal_total = group[bandar_lokal_columns].sum().sum() * closing_price
            bandar_asing_total = group[bandar_asing_columns].sum().sum() * closing_price
            big_inv_lokal_total = group[big_investor_lokal_columns].sum().sum() * closing_price
            big_inv_asing_total = group[big_investor_asing_columns].sum().sum() * closing_price

            response += f"\n📅 {month.strftime('%b %Y')}\n"
            response += f"💸 Ritel: Rp {ritel_total:,.0f}\n"
            response += f"🏦 Bandar Lokal: Rp {bandar_lokal_total:,.0f}\n"
            response += f"🌏 Bandar Asing: Rp {bandar_asing_total:,.0f}\n"
            response += f"💼 Big Investor Lokal: Rp {big_inv_lokal_total:,.0f}\n"
            response += f"💼 Big Investor Asing: Rp {big_inv_asing_total:,.0f}\n"
            response += "─" * 30 + "\n"

        bot.reply_to(message, response)

        viewer.combined_df = None

    except Exception as e:
        bot.reply_to(message, f"❌ Error: {str(e)}")
        logger.error(f"Error in /hol command: {e}")

    finally:
        viewer.combined_df = None
        import gc
        gc.collect()

@bot.callback_query_handler(func=lambda call: call.data.startswith('wl_page_'))
def handle_watchlist_pagination(call):
    try:
        parts = call.data.split('_', 3)  # Maks 4 bagian
        cap_filter = parts[2]
        page_str = parts[3]
        page = int(page_str)

        filter_param = None if cap_filter == 'all' else cap_filter
        stocks = viewer.get_watchlist_stocks(filter_param)
        response = viewer.format_watchlist_response(stocks, filter_param)

        # Split jadi halaman-halaman
        page_size = 4000
        pages = [response[i:i + page_size] for i in range(0, len(response), page_size)]

        if page >= len(pages):
            bot.answer_callback_query(call.id, "❌ Halaman tidak ditemukan.")
            return

        # Update teks & tombol
        bot.edit_message_text(
            pages[page],
            chat_id=call.message.chat.id,
            message_id=call.message.message_id
        )

        # Buat tombol prev/next
        markup = types.InlineKeyboardMarkup()
        if page > 0:
            markup.add(types.InlineKeyboardButton("⬅️ Prev", callback_data=f"wl_page_{cap_filter}_{page-1}"))
        if page < len(pages)-1:
            markup.add(types.InlineKeyboardButton("➡️ Next", callback_data=f"wl_page_{cap_filter}_{page+1}"))
        bot.edit_message_reply_markup(
            chat_id=call.message.chat.id,
            message_id=call.message.message_id,
            reply_markup=markup
        )
        bot.answer_callback_query(call.id)

    except Exception as e:
        bot.answer_callback_query(call.id, f"❌ Error: {str(e)}")
        logger.error(f"Error in pagination: {e}")

@bot.callback_query_handler(func=lambda call: call.data.startswith('wl_'))
def handle_watchlist_filter(call):
    cap_filter = call.data[3:]  # Remove 'wl_' prefix

    try:
        bot.answer_callback_query(call.id, "🔍 Menganalisa watchlist...")  
        filter_param = None if cap_filter == 'all' else cap_filter
        stocks = viewer.get_watchlist_stocks(filter_param)
        response = viewer.format_watchlist_response(stocks, filter_param)
        bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)

        # Split response jadi halaman-halaman (maks 4000 karakter per halaman)
        page_size = 4000
        pages = [response[i:i + page_size] for i in range(0, len(response), page_size)]

        if not pages:
            bot.send_message(call.message.chat.id, "❌ Tidak ada data.")
            return

        # Kirim halaman pertama + tombol next jika ada halaman berikutnya
        msg = bot.send_message(call.message.chat.id, pages[0])

        if len(pages) > 1:
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("📄 Next (2)", callback_data=f"wl_page_{cap_filter}_1"))
            bot.edit_message_reply_markup(chat_id=msg.chat.id, message_id=msg.message_id, reply_markup=markup)

    except Exception as e:
        bot.answer_callback_query(call.id, f"❌ Error: {str(e)}")
        logger.error(f"Error in watchlist filter: {e}")
        
    
if __name__ == "__main__":
    print("🤖 Bot started successfully!")
    print(f"📁 Data folder: {viewer.data_folder}")
    
    data_info = viewer.get_data_info()
    if isinstance(data_info['regular'], str):
        print("❌ No data loaded - make sure Excel files are in the 'data' folder")
    else:
        regular = data_info['regular']
        print(f"✅ Loaded {regular['total_records']} records from {regular['unique_codes']} stocks")
        print(f"📅 Date range: {regular['date_range']}")

    print("📝 Remember to replace BOT_TOKEN with your actual bot token")
    bot.infinity_polling()