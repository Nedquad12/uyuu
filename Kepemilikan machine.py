import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime
import os
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class MultiFileStockDataViewer:
    def __init__(self, root):
        self.root = root
        self.root.title("Multi-File Stock Data Viewer")
        self.root.geometry("1200x800")
        
        # Data storage
        self.combined_df = None
        self.selected_files = []
        
        # Create GUI
        self.create_widgets()
        
    def create_widgets(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # File selection frame
        file_frame = ttk.LabelFrame(main_frame, text="File Selection (Max 6 files)", padding="10")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Buttons frame
        buttons_frame = ttk.Frame(file_frame)
        buttons_frame.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        ttk.Button(buttons_frame, text="Add Excel Files", command=self.add_files).grid(row=0, column=0, padx=(0, 10))
        ttk.Button(buttons_frame, text="Clear Files", command=self.clear_files).grid(row=0, column=1, padx=(0, 10))
        ttk.Button(buttons_frame, text="Load Data", command=self.load_data).grid(row=0, column=2)
        
        # Files listbox
        self.files_listbox = tk.Listbox(file_frame, height=4)
        self.files_listbox.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        # Search frame
        search_frame = ttk.LabelFrame(main_frame, text="Search & Export", padding="10")
        search_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Search controls
        search_controls = ttk.Frame(search_frame)
        search_controls.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        ttk.Label(search_controls, text="Stock Code:").grid(row=0, column=0, padx=(0, 10))
        self.code_entry = ttk.Entry(search_controls, width=15)
        self.code_entry.grid(row=0, column=1, padx=(0, 10))
        self.code_entry.bind('<Return>', lambda e: self.search_stock())
        
        ttk.Button(search_controls, text="Search", command=self.search_stock).grid(row=0, column=2, padx=(0, 20))
        ttk.Button(search_controls, text="Export to Excel", command=self.export_to_excel).grid(row=0, column=3)
        
        # Results frame with treeview
        results_frame = ttk.LabelFrame(main_frame, text="Results", padding="10")
        results_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Create treeview for tabular display
        self.create_treeview(results_frame)
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("Ready - Please add Excel files")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        ttk.Button(main_frame, text="Show Line Chart", command=self.open_line_chart_window).grid(row=5, column=0, pady=5)

        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        file_frame.columnconfigure(0, weight=1)
        search_frame.columnconfigure(0, weight=1)
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
    
    def create_treeview(self, parent):
        # Frame for treeview and scrollbars
        tree_frame = ttk.Frame(parent)
        tree_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        
        # Create treeview
        columns = ['Date', 'Code', 'Type', 'Price', 'Local IS', 'Local CP', 'Local PF', 'Local IB', 
                  'Local ID', 'Local MF', 'Local SC', 'Foreign IS', 'Foreign CP', 'Foreign PF', 
                  'Foreign IB', 'Foreign ID', 'Foreign MF', 'Foreign SC', 'Total Local', 'Total Foreign']
        
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        
        # Configure column headings and widths
        for col in columns:
            self.tree.heading(col, text=col)
            if col in ['Date', 'Code', 'Type']:
                self.tree.column(col, width=100)
            elif col in ['Total Local', 'Total Foreign']:
                self.tree.column(col, width=120)
            else:
                self.tree.column(col, width=80)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Grid layout
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
    
    def add_files(self):
        if len(self.selected_files) >= 6:
            messagebox.showwarning("Warning", "Maximum 6 files allowed!")
            return
        
        remaining_slots = 6 - len(self.selected_files)
        file_paths = filedialog.askopenfilenames(
            title=f"Select Excel Files (Max {remaining_slots} more)",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        for file_path in file_paths:
            if len(self.selected_files) < 6 and file_path not in self.selected_files:
                self.selected_files.append(file_path)
                self.files_listbox.insert(tk.END, os.path.basename(file_path))
        
        self.status_var.set(f"Selected {len(self.selected_files)}/6 files")
    
    def clear_files(self):
        self.selected_files.clear()
        self.files_listbox.delete(0, tk.END)
        self.combined_df = None
        self.clear_treeview()
        self.status_var.set("Files cleared - Please add Excel files")
    
    def load_data(self):
        if not self.selected_files:
            messagebox.showwarning("Warning", "Please select at least one Excel file!")
            return
        
        try:
            dfs = []
            for file_path in self.selected_files:
                df = pd.read_excel(file_path)
                dfs.append(df)
            
            # Combine all dataframes
            self.combined_df = pd.concat(dfs, ignore_index=True)
            
            # Convert Date column
            self.combined_df['Date'] = pd.to_datetime(self.combined_df['Date'])
            
            # Sort by date (newest first)
            self.combined_df = self.combined_df.sort_values('Date', ascending=False)
            
            self.status_var.set(f"Loaded {len(self.combined_df)} records from {len(self.selected_files)} files")
            messagebox.showinfo("Success", f"Successfully loaded data from {len(self.selected_files)} files!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load files: {str(e)}")
    
    def clear_treeview(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
    
    def search_stock(self):
        if self.combined_df is None:
            messagebox.showwarning("Warning", "Please load data first!")
            return
        
        code = self.code_entry.get().upper().strip()
        if not code:
            messagebox.showwarning("Warning", "Please enter a stock code!")
            return
        
        # Filter data by code and limit to 6 records
        stock_data = self.combined_df[self.combined_df['Code'].str.upper() == code].head(6)
        
        if stock_data.empty:
            self.clear_treeview()
            self.status_var.set(f"No data found for stock code: {code}")
            return
        
        # Display in treeview
        self.display_in_treeview(stock_data)
        self.status_var.set(f"Found {len(stock_data)} records for {code}")
    
    def display_in_treeview(self, data):
        self.clear_treeview()
        
        local_columns = ['Local IS', 'Local CP', 'Local PF', 'Local IB', 'Local ID', 'Local MF', 'Local SC']
        foreign_columns = ['Foreign IS', 'Foreign CP', 'Foreign PF', 'Foreign IB', 'Foreign ID', 'Foreign MF', 'Foreign SC']
        
        for _, row in data.iterrows():
            # Calculate totals
            local_total = sum(row.get(col, 0) for col in local_columns if pd.notna(row.get(col, 0)))
            foreign_total = sum(row.get(col, 0) for col in foreign_columns if pd.notna(row.get(col, 0)))
            
            # Format values
            values = [
                row['Date'].strftime('%d-%b-%Y'),
                row.get('Code', ''),
                row.get('Type', ''),
                f"{row.get('Price', 0):,.0f}",
                f"{row.get('Local IS', 0):,.0f}",
                f"{row.get('Local CP', 0):,.0f}",
                f"{row.get('Local PF', 0):,.0f}",
                f"{row.get('Local IB', 0):,.0f}",
                f"{row.get('Local ID', 0):,.0f}",
                f"{row.get('Local MF', 0):,.0f}",
                f"{row.get('Local SC', 0):,.0f}",
                f"{row.get('Foreign IS', 0):,.0f}",
                f"{row.get('Foreign CP', 0):,.0f}",
                f"{row.get('Foreign PF', 0):,.0f}",
                f"{row.get('Foreign IB', 0):,.0f}",
                f"{row.get('Foreign ID', 0):,.0f}",
                f"{row.get('Foreign MF', 0):,.0f}",
                f"{row.get('Foreign SC', 0):,.0f}",
                f"{local_total:,.0f}",
                f"{foreign_total:,.0f}"
            ]
            
            self.tree.insert('', 'end', values=values)
    
    def export_to_excel(self):
        if self.combined_df is None:
            messagebox.showwarning("Warning", "Please load data first!")
            return
        
        code = self.code_entry.get().upper().strip()
        if not code:
            messagebox.showwarning("Warning", "Please enter a stock code to export!")
            return
        
        # Filter data
        stock_data = self.combined_df[self.combined_df['Code'].str.upper() == code].head(6)
        
        if stock_data.empty:
            messagebox.showwarning("Warning", f"No data found for stock code: {code}")
            return
        
        # Save file dialog
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialname=f"{code}_analysis.xlsx"
        )
        
        if not file_path:
            return
        
        try:
            self.create_excel_report(stock_data, file_path, code)
            messagebox.showinfo("Success", f"Data exported successfully to:\n{file_path}")
            self.status_var.set(f"Exported {len(stock_data)} records for {code}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export data: {str(e)}")
    
    def create_excel_report(self, data, file_path, code):
        wb = Workbook()
        ws = wb.active
        ws.title = f"{code} Analysis"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center')
        
        # Headers
        headers = ['Date', 'Code', 'Type', 'Price', 'Local IS', 'Local CP', 'Local PF', 'Local IB', 
                  'Local ID', 'Local MF', 'Local SC', 'Foreign IS', 'Foreign CP', 'Foreign PF', 
                  'Foreign IB', 'Foreign ID', 'Foreign MF', 'Foreign SC']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
        
        # Add Total columns headers
        ws.cell(row=1, column=len(headers)+1, value="Total Local").font = header_font
        ws.cell(row=1, column=len(headers)+1).fill = header_fill
        ws.cell(row=1, column=len(headers)+1).border = border
        ws.cell(row=1, column=len(headers)+1).alignment = center_align
        
        ws.cell(row=1, column=len(headers)+2, value="Total Foreign").font = header_font
        ws.cell(row=1, column=len(headers)+2).fill = header_fill
        ws.cell(row=1, column=len(headers)+2).border = border
        ws.cell(row=1, column=len(headers)+2).alignment = center_align
        
        # Write data
        local_cols = ['Local IS', 'Local CP', 'Local PF', 'Local IB', 'Local ID', 'Local MF', 'Local SC']
        foreign_cols = ['Foreign IS', 'Foreign CP', 'Foreign PF', 'Foreign IB', 'Foreign ID', 'Foreign MF', 'Foreign SC']
        
        for row_idx, (_, row_data) in enumerate(data.iterrows(), 2):
            # Basic data
            ws.cell(row=row_idx, column=1, value=row_data['Date'].strftime('%d-%b-%Y')).border = border
            ws.cell(row=row_idx, column=2, value=row_data.get('Code', '')).border = border
            ws.cell(row=row_idx, column=3, value=row_data.get('Type', '')).border = border
            ws.cell(row=row_idx, column=4, value=row_data.get('Price', 0)).border = border
            
            # Local and Foreign data
            for col_idx, header in enumerate(headers[4:], 5):
                value = row_data.get(header, 0)
                ws.cell(row=row_idx, column=col_idx, value=value).border = border
            
            # Add formulas for totals
            local_range = f"E{row_idx}:K{row_idx}"  # Local columns
            foreign_range = f"L{row_idx}:R{row_idx}"  # Foreign columns
            
            total_local_cell = ws.cell(row=row_idx, column=len(headers)+1)
            total_local_cell.value = f"=SUM({local_range})"
            total_local_cell.border = border
            
            total_foreign_cell = ws.cell(row=row_idx, column=len(headers)+2)
            total_foreign_cell.value = f"=SUM({foreign_range})"
            total_foreign_cell.border = border
        
        # Add summary section
        summary_row = len(data) + 3
        ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=14)
        ws.cell(row=summary_row+1, column=1, value=f"Stock Code: {code}")
        ws.cell(row=summary_row+2, column=1, value=f"Total Records: {len(data)}")
        ws.cell(row=summary_row+3, column=1, value=f"Date Range: {data['Date'].min().strftime('%d-%b-%Y')} to {data['Date'].max().strftime('%d-%b-%Y')}")
        
        # Add grand total formulas
        last_row = len(data) + 1
        ws.cell(row=summary_row+5, column=1, value="Grand Total Local:").font = Font(bold=True)
        ws.cell(row=summary_row+5, column=2, value=f"=SUM(S2:S{last_row})").font = Font(bold=True)
        ws.cell(row=summary_row+6, column=1, value="Grand Total Foreign:").font = Font(bold=True)
        ws.cell(row=summary_row+6, column=2, value=f"=SUM(T2:T{last_row})").font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
  
    def open_line_chart_window(self):
         if self.combined_df is None:
            messagebox.showwarning("Warning", "Please load data first!")
            return

         window = tk.Toplevel(self.root)
         window.title("Line Chart Viewer")
         window.geometry("400x600")

         ttk.Label(window, text="Select columns to plot:", font=("Arial", 12)).pack(pady=10)


    # Available fields
         self.plot_fields = [
              'Local IS', 'Local CP', 'Local PF', 'Local IB', 'Local ID', 'Local MF', 'Local SC',
               'Foreign IS', 'Foreign CP', 'Foreign PF', 'Foreign IB', 'Foreign ID', 'Foreign MF', 'Foreign SC'
            ]
         self.check_vars = {}
    
         checkbox_frame = ttk.Frame(window)
         checkbox_frame.pack(pady=5)
    
         for field in self.plot_fields:
            var = tk.BooleanVar()
            chk = ttk.Checkbutton(checkbox_frame, text=field, variable=var)
            chk.pack(anchor='w')
            self.check_vars[field] = var

    # Total Local & Foreign
         self.total_local_var = tk.BooleanVar()
         self.total_foreign_var = tk.BooleanVar()

         ttk.Checkbutton(window, text="Total Local", variable=self.total_local_var).pack(anchor='w', padx=10, pady=(10, 0))
         ttk.Checkbutton(window, text="Total Foreign", variable=self.total_foreign_var).pack(anchor='w', padx=10)

    # Select All
         self.select_all_var = tk.BooleanVar()
         def toggle_all():
             for var in self.check_vars.values():
                 var.set(self.select_all_var.get())
         ttk.Checkbutton(window, text="Select All", variable=self.select_all_var, command=toggle_all).pack(anchor='w', padx=10, pady=(10, 5))

    # Plot button
         ttk.Button(window, text="Plot Chart", command=self.plot_line_chart).pack(pady=20)

    def plot_line_chart(self):
        if self.combined_df is None:
           messagebox.showwarning("Warning", "Please load data first!")
           return

        selected_fields = [field for field, var in self.check_vars.items() if var.get()]
    
    # Include total columns if checked
        if self.total_local_var.get():
           self.combined_df['Total Local'] = self.combined_df[[f for f in self.plot_fields if f.startswith('Local')]].sum(axis=1)
           selected_fields.append('Total Local')
        if self.total_foreign_var.get():
           self.combined_df['Total Foreign'] = self.combined_df[[f for f in self.plot_fields if f.startswith('Foreign')]].sum(axis=1)
           selected_fields.append('Total Foreign')

        if not selected_fields:
           messagebox.showwarning("Warning", "Please select at least one field to plot!")
           return

    # Filter stock if applicable
        code = self.code_entry.get().upper().strip()
        df = self.combined_df.copy()
        if code:
           df = df[df['Code'].str.upper() == code]
           if df.empty:
               messagebox.showinfo("No Data", f"No data found for stock code: {code}")
               return

    # Group by date and sum
        grouped = df.groupby('Date')[selected_fields].sum().sort_index()

        plt.figure(figsize=(10, 6))
        for field in selected_fields:
            plt.plot(grouped.index, grouped[field], marker='o', label=field)

        plt.xlabel('Date')
        plt.ylabel('Value')
        plt.title(f"Line Chart{' for ' + code if code else ''}")
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        plt.show()



def main():
    root = tk.Tk()
    app = MultiFileStockDataViewer(root)
    root.mainloop()

if __name__ == "__main__":
    main()
